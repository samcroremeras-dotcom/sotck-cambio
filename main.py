import os
import io
import uuid
import requests
import json
from datetime import datetime, timedelta
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from psycopg2.extras import RealDictCursor
import psycopg2
import openpyxl

app = FastAPI()

DATABASE_URL = os.getenv("DATABASE_URL")
TN_CLIENT_ID = os.getenv("TN_CLIENT_ID")
TN_CLIENT_SECRET = os.getenv("TN_CLIENT_SECRET")
TN_ACCESS_TOKEN = os.getenv("TN_ACCESS_TOKEN")
TN_STORE_ID = os.getenv("TN_STORE_ID")
TN_ENVIO_PRODUCT_ID = os.getenv("TN_ENVIO_PRODUCT_ID")
TN_ENVIO_VARIANT_ID = os.getenv("TN_ENVIO_VARIANT_ID")
TN_ENVIO_PRECIO = os.getenv("TN_ENVIO_PRECIO", "0")

if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

def get_conn():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def init_db():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS stock (
                    id SERIAL PRIMARY KEY,
                    nombre TEXT,
                    categoria TEXT,
                    talle TEXT,
                    color TEXT,
                    cantidad INTEGER DEFAULT 0,
                    imagen_url TEXT,
                    link_tienda TEXT,
                    creado_en TIMESTAMP DEFAULT NOW()
                );
                CREATE TABLE IF NOT EXISTS tokens_cambio (
                    token_id TEXT PRIMARY KEY,
                    orden_nro TEXT,
                    expira_at TIMESTAMP,
                    usado BOOLEAN DEFAULT FALSE,
                    remera_elegida_id INTEGER REFERENCES stock(id)
                );
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS cliente_email TEXT;
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS cliente_nombre TEXT;
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS productos_originales JSONB;
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS finalizado BOOLEAN DEFAULT FALSE;
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS creado_en TIMESTAMP DEFAULT NOW();
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS cliente_paga_envio BOOLEAN DEFAULT FALSE;
                ALTER TABLE tokens_cambio ADD COLUMN IF NOT EXISTS cliente_id BIGINT;
                ALTER TABLE cambios ADD COLUMN IF NOT EXISTS orden_envio_tn_id BIGINT;
                ALTER TABLE cambios ADD COLUMN IF NOT EXISTS orden_envio_tn_number TEXT;

                CREATE TABLE IF NOT EXISTS cambios (
                    id SERIAL PRIMARY KEY,
                    token_id TEXT REFERENCES tokens_cambio(token_id),
                    orden_nro TEXT,
                    cliente_email TEXT,
                    producto_original JSONB,
                    remera_elegida_id INTEGER REFERENCES stock(id),
                    remera_elegida_nombre TEXT,
                    remera_elegida_talle TEXT,
                    remera_elegida_color TEXT,
                    remera_elegida_imagen TEXT,
                    estado TEXT DEFAULT 'pendiente_recepcion',
                    motivo_rechazo TEXT,
                    aprobado_por TEXT,
                    aprobado_en TIMESTAMP,
                    creado_en TIMESTAMP DEFAULT NOW(),
                    actualizado_en TIMESTAMP DEFAULT NOW()
                );

                CREATE TABLE IF NOT EXISTS cambios_historial (
                    id SERIAL PRIMARY KEY,
                    cambio_id INTEGER REFERENCES cambios(id) ON DELETE CASCADE,
                    token_id TEXT,
                    accion TEXT,
                    datos JSONB,
                    creado_en TIMESTAMP DEFAULT NOW()
                );

                CREATE INDEX IF NOT EXISTS idx_cambios_token ON cambios(token_id);
                CREATE INDEX IF NOT EXISTS idx_cambios_estado ON cambios(estado);
            """)
            conn.commit()

init_db()

@app.get("/health")
def health():
    return {"status": "ok"}

@app.get("/auth/callback")
def auth_callback(code: str):
    response = requests.post(
        "https://www.tiendanube.com/apps/authorize/token",
        json={
            "client_id": TN_CLIENT_ID,
            "client_secret": TN_CLIENT_SECRET,
            "grant_type": "authorization_code",
            "code": code
        },
        headers={"User-Agent": "Samcro Stock (samcroremeras@gmail.com)"}
    )
    if response.status_code == 200:
        data = response.json()
        return {
            "ok": True,
            "TN_ACCESS_TOKEN": data.get("access_token"),
            "TN_STORE_ID": data.get("user_id"),
            "instruccion": "Copia estos dos valores y agregalos como variables en Railway"
        }
    return {"ok": False, "detalle": response.text}

class Remera(BaseModel):
    nombre: str
    categoria: str = ""
    talle: str
    color: str = ""
    cantidad: int = 0
    imagen_url: str = ""
    link_tienda: str = ""

@app.get("/api/stock")
def obtener_stock():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM stock ORDER BY id DESC;")
            return cur.fetchall()

@app.post("/api/eliminar-ultima-imagen")
def eliminar_ultima_imagen(data: dict):
    product_ids = data.get("product_ids", [])
    ok = []
    errores = []
    for pid in product_ids:
        # Traer imágenes del producto
        res = requests.get(
            f"https://api.tiendanube.com/v1/{TN_STORE_ID}/products/{pid}/images",
            headers={
                "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
            }
        )
        if res.status_code != 200 or not res.json():
            errores.append({"id": pid, "error": "no se pudieron traer imagenes"})
            continue
        imagenes = res.json()
        ultima = imagenes[-1]
        # Borrar la última
        res2 = requests.delete(
            f"https://api.tiendanube.com/v1/{TN_STORE_ID}/products/{pid}/images/{ultima['id']}",
            headers={
                "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
            }
        )
        if res2.status_code in (200, 201, 204):
            ok.append(pid)
        else:
            errores.append({"id": pid, "error": res2.text})
    return {"ok": len(ok), "errores": errores}
    
@app.post("/api/stock")
def agregar_remera(r: Remera):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO stock (nombre, categoria, talle, color, cantidad, imagen_url, link_tienda)
                VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id;
            """, (r.nombre, r.categoria, r.talle, r.color, r.cantidad, r.imagen_url, r.link_tienda))
            nuevo_id = cur.fetchone()["id"]
            conn.commit()
            return {"ok": True, "id": nuevo_id}

@app.put("/api/stock/{id}")
def editar_remera(id: int, r: Remera):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE stock SET nombre=%s, categoria=%s, talle=%s, color=%s,
                cantidad=%s, imagen_url=%s, link_tienda=%s WHERE id=%s;
            """, (r.nombre, r.categoria, r.talle, r.color, r.cantidad, r.imagen_url, r.link_tienda, id))
            conn.commit()
            return {"ok": True}

@app.delete("/api/stock/{id}")
def eliminar_remera(id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM stock WHERE id=%s;", (id,))
            conn.commit()
            return {"ok": True}

@app.post("/api/importar-excel")
async def importar_excel(file: UploadFile = File(...)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    total = 0
    por_hoja = {}
    with get_conn() as conn:
        with conn.cursor() as cur:
            for ws in wb.worksheets:
                if ws.max_row < 2:
                    continue
                headers = [str(c.value).lower().strip() if c.value is not None else "" for c in ws[1]]
                if "nombre" not in headers:
                    por_hoja[ws.title] = 0
                    continue
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    if not data.get("nombre"):
                        continue
                    cur.execute("""
                        INSERT INTO stock (nombre, categoria, talle, color, cantidad, imagen_url, link_tienda)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (
                        str(data.get("nombre", "")),
                        str(data.get("categoria", "") or ws.title),
                        str(data.get("talle", "")),
                        str(data.get("color", "")),
                        int(data.get("cantidad") or 0),
                        str(data.get("imagen_url", "")),
                        str(data.get("link_tienda", ""))
                    ))
                    count += 1
                por_hoja[ws.title] = count
                total += count
            conn.commit()
    return {"ok": True, "importadas": total, "hojas": por_hoja}


@app.post("/api/stock/vaciar")
async def vaciar_stock(confirmacion: str = ""):
    if confirmacion != "BORRAR TODO EL STOCK":
        raise HTTPException(status_code=400, detail="Confirmación inválida")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS n FROM stock;")
            antes = cur.fetchone()["n"]
            cur.execute("DELETE FROM stock;")
            conn.commit()
    return {"ok": True, "borradas": antes}

@app.get("/api/exportar-excel")
def exportar_excel():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT nombre, categoria, talle, color, cantidad, imagen_url, link_tienda FROM stock ORDER BY id DESC;")
            rows = cur.fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["nombre", "categoria", "talle", "color", "cantidad", "imagen_url", "link_tienda"])
    for r in rows:
        ws.append([r["nombre"], r["categoria"], r["talle"], r["color"], r["cantidad"], r["imagen_url"], r["link_tienda"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock.xlsx"})

def _tn_buscar_orden(orden_nro: str):
    """Busca una orden en Tienda Nube por su numero. Devuelve dict o None."""
    try:
        res = requests.get(
            f"https://api.tiendanube.com/v1/{TN_STORE_ID}/orders",
            headers={
                "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
            },
            params={"q": orden_nro, "per_page": 5}
        )
        if res.status_code != 200:
            return None
        ordenes = res.json()
        for o in ordenes:
            if str(o.get("number")) == str(orden_nro):
                return o
        return ordenes[0] if ordenes else None
    except Exception:
        return None


def _tn_crear_orden_envio(cliente_id, cliente_email, cliente_nombre, orden_nro_origen):
    """Crea una orden de venta en TN con el item ENVIO CORREO ARGENTINO."""
    if not TN_ENVIO_PRODUCT_ID:
        return None, "TN_ENVIO_PRODUCT_ID no configurado"
    payload = {
        "contact_email": cliente_email,
        "contact_name": cliente_nombre,
        "products": [{
            "product_id": int(TN_ENVIO_PRODUCT_ID),
            "variant_id": int(TN_ENVIO_VARIANT_ID) if TN_ENVIO_VARIANT_ID else None,
            "quantity": 1,
            "price": str(TN_ENVIO_PRECIO)
        }],
        "note": f"Envio cambio - orden original #{orden_nro_origen}",
        "send_confirmation_email": True,
        "send_fulfillment_email": False
    }
    if cliente_id:
        payload["customer_id"] = int(cliente_id)
    payload["products"] = [{k: v for k, v in p.items() if v is not None} for p in payload["products"]]
    try:
        res = requests.post(
            f"https://api.tiendanube.com/v1/{TN_STORE_ID}/orders",
            headers={
                "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                "User-Agent": "Samcro Stock (samcroremeras@gmail.com)",
                "Content-Type": "application/json"
            },
            json=payload,
            timeout=20
        )
        if res.status_code not in (200, 201):
            return None, f"TN respondio {res.status_code}: {res.text[:200]}"
        data = res.json()
        return data, None
    except Exception as e:
        return None, str(e)


@app.post("/api/tokens")
def crear_token(orden_nro: str, cliente_paga_envio: bool = False):
    orden = _tn_buscar_orden(orden_nro)
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada en Tienda Nube")

    cliente = orden.get("customer") or {}
    cliente_email = (cliente.get("email") or "").lower().strip()
    cliente_nombre = cliente.get("name") or ""
    cliente_id = cliente.get("id")

    productos_originales = []
    for prod in (orden.get("products") or []):
        productos_originales.append({
            "id": prod.get("id"),
            "product_id": prod.get("product_id"),
            "variant_id": prod.get("variant_id"),
            "nombre": prod.get("name") or "",
            "talle": ((prod.get("variant_values") or [None])[0] if prod.get("variant_values") else None) or "",
            "cantidad": prod.get("quantity") or 1,
            "imagen": (prod.get("image") or {}).get("src") if isinstance(prod.get("image"), dict) else "",
            "precio": prod.get("price")
        })

    token = str(uuid.uuid4())[:8]
    expira = datetime.now() + timedelta(days=5)

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO tokens_cambio
                  (token_id, orden_nro, expira_at, cliente_email, cliente_nombre, cliente_id,
                   productos_originales, finalizado, cliente_paga_envio)
                VALUES (%s, %s, %s, %s, %s, %s, %s, FALSE, %s);
            """, (token, str(orden_nro), expira, cliente_email, cliente_nombre, cliente_id,
                  json.dumps(productos_originales), bool(cliente_paga_envio)))
            conn.commit()

    return {
        "token": token,
        "link": f"https://samcro-stock-production.up.railway.app/cambios/{token}",
        "cliente_email": cliente_email,
        "cliente_nombre": cliente_nombre,
        "productos_originales": productos_originales,
        "expira_at": expira.isoformat(),
        "cliente_paga_envio": bool(cliente_paga_envio)
    }


class ValidarAccesoPayload(BaseModel):
    token: str
    email: str

@app.post("/api/validar-acceso")
def validar_acceso(payload: ValidarAccesoPayload):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM tokens_cambio WHERE token_id=%s;", (payload.token,))
            t = cur.fetchone()
    if not t:
        raise HTTPException(status_code=404, detail="Link invalido")
    if datetime.now() > t["expira_at"]:
        raise HTTPException(status_code=410, detail="Link expirado")
    if t.get("finalizado"):
        raise HTTPException(status_code=409, detail="Cambio ya finalizado")
    if (t.get("cliente_email") or "").lower().strip() != payload.email.lower().strip():
        raise HTTPException(status_code=403, detail="Email no coincide con la orden")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, producto_original, remera_elegida_id, remera_elegida_nombre,
                       remera_elegida_talle, remera_elegida_color, remera_elegida_imagen, estado
                FROM cambios
                WHERE token_id=%s
                ORDER BY creado_en ASC;
            """, (payload.token,))
            cambios = cur.fetchall()

    return {
        "ok": True,
        "orden_nro": t["orden_nro"],
        "cliente_nombre": t.get("cliente_nombre") or "",
        "productos_originales": t.get("productos_originales") or [],
        "selecciones_previas": cambios,
        "expira_at": t["expira_at"].isoformat()
    }


class SeleccionItem(BaseModel):
    producto_original_index: int
    remera_id: int

class GuardarSeleccionPayload(BaseModel):
    token: str
    email: str
    selecciones: list[SeleccionItem]
    finalizar: bool = False

@app.post("/api/cambios/seleccionar")
def guardar_seleccion(payload: GuardarSeleccionPayload):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM tokens_cambio WHERE token_id=%s;", (payload.token,))
            t = cur.fetchone()
            if not t:
                raise HTTPException(status_code=404, detail="Link invalido")
            if datetime.now() > t["expira_at"]:
                raise HTTPException(status_code=410, detail="Link expirado")
            if t.get("finalizado"):
                raise HTTPException(status_code=409, detail="Cambio ya finalizado")
            if (t.get("cliente_email") or "").lower().strip() != payload.email.lower().strip():
                raise HTTPException(status_code=403, detail="Email no coincide")

            productos_originales = t.get("productos_originales") or []

            # Borrar selecciones previas no aprobadas y volver a crear
            cur.execute("""
                DELETE FROM cambios
                WHERE token_id=%s AND estado IN ('pendiente_recepcion','pendiente_aprobacion');
            """, (payload.token,))

            cambios_creados = []
            for s in payload.selecciones:
                if s.producto_original_index < 0 or s.producto_original_index >= len(productos_originales):
                    continue
                prod_orig = productos_originales[s.producto_original_index]
                cur.execute("""
                    SELECT s.id, s.nombre, s.talle, s.color, s.imagen_url,
                           s.cantidad - COALESCE((
                               SELECT COUNT(*) FROM cambios
                               WHERE remera_elegida_id = s.id
                                 AND estado IN ('pendiente_recepcion','pendiente_aprobacion')
                           ), 0) AS disponible
                    FROM stock s WHERE s.id=%s;
                """, (s.remera_id,))
                stock = cur.fetchone()
                if not stock or (stock.get("disponible") or 0) <= 0:
                    raise HTTPException(status_code=409, detail=f"La remera {stock['nombre'] if stock else ''} talle {stock['talle'] if stock else ''} ya fue reservada por otro cliente. Volve a elegir.")
                cur.execute("""
                    INSERT INTO cambios
                      (token_id, orden_nro, cliente_email, producto_original,
                       remera_elegida_id, remera_elegida_nombre, remera_elegida_talle,
                       remera_elegida_color, remera_elegida_imagen, estado)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'pendiente_recepcion')
                    RETURNING id;
                """, (
                    payload.token, t["orden_nro"], t["cliente_email"], json.dumps(prod_orig),
                    stock["id"], stock["nombre"], stock["talle"],
                    stock["color"], stock["imagen_url"]
                ))
                cid = cur.fetchone()["id"]
                cur.execute("""
                    INSERT INTO cambios_historial (cambio_id, token_id, accion, datos)
                    VALUES (%s,%s,'seleccion_creada',%s);
                """, (cid, payload.token, json.dumps({
                    "remera_id": stock["id"],
                    "remera_nombre": stock["nombre"],
                    "talle": stock["talle"],
                    "color": stock["color"]
                })))
                cambios_creados.append(cid)

            if payload.finalizar:
                cur.execute("UPDATE tokens_cambio SET finalizado=TRUE WHERE token_id=%s;", (payload.token,))

            conn.commit()
    return {"ok": True, "cambios_ids": cambios_creados, "finalizado": payload.finalizar}


@app.get("/api/cambios/pendientes")
def listar_cambios_pendientes(estado: str = ""):
    with get_conn() as conn:
        with conn.cursor() as cur:
            base = """
                SELECT c.*, COALESCE(t.cliente_paga_envio, FALSE) AS cliente_paga_envio
                FROM cambios c
                LEFT JOIN tokens_cambio t ON t.token_id = c.token_id
            """
            if estado:
                cur.execute(base + " WHERE c.estado=%s ORDER BY c.creado_en DESC;", (estado,))
            else:
                cur.execute(base + " WHERE c.estado IN ('pendiente_recepcion','pendiente_aprobacion') ORDER BY c.creado_en DESC;")
            return cur.fetchall()


class AprobarPayload(BaseModel):
    aprobado_por: str = "admin"

@app.post("/api/cambios/{cambio_id}/aprobar")
def aprobar_cambio(cambio_id: int, payload: AprobarPayload):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM cambios WHERE id=%s;", (cambio_id,))
            c = cur.fetchone()
            if not c:
                raise HTTPException(status_code=404, detail="Cambio no encontrado")
            if c["estado"] == "aprobado":
                return {"ok": True, "ya_aprobado": True}

            # Descontar stock de la remera nueva
            cur.execute(
                "UPDATE stock SET cantidad = GREATEST(cantidad - 1, 0) WHERE id=%s;",
                (c["remera_elegida_id"],)
            )
            # Sumar stock de la remera devuelta si la encontramos en stock por nombre+talle
            prod_orig = c.get("producto_original") or {}
            nombre_orig = (prod_orig.get("nombre") or "").strip()
            talle_orig = (prod_orig.get("talle") or "").strip()
            if nombre_orig and talle_orig:
                cur.execute("""
                    UPDATE stock SET cantidad = cantidad + 1
                    WHERE LOWER(TRIM(nombre)) = LOWER(%s) AND LOWER(TRIM(talle)) = LOWER(%s);
                """, (nombre_orig, talle_orig))

            cur.execute("""
                UPDATE cambios
                SET estado='aprobado', aprobado_por=%s, aprobado_en=NOW(), actualizado_en=NOW()
                WHERE id=%s;
            """, (payload.aprobado_por, cambio_id))

            cur.execute("""
                INSERT INTO cambios_historial (cambio_id, token_id, accion, datos)
                VALUES (%s,%s,'aprobado',%s);
            """, (cambio_id, c["token_id"], json.dumps({"aprobado_por": payload.aprobado_por})))

            # Si el cliente paga el envio, crear la orden TN con el item ENVIO CORREO ARGENTINO
            cur.execute("""
                SELECT cliente_paga_envio, cliente_id, cliente_email, cliente_nombre, orden_nro
                FROM tokens_cambio WHERE token_id=%s;
            """, (c["token_id"],))
            t = cur.fetchone() or {}
            envio_info = None
            if t.get("cliente_paga_envio"):
                data, err = _tn_crear_orden_envio(
                    t.get("cliente_id"),
                    t.get("cliente_email") or c.get("cliente_email"),
                    t.get("cliente_nombre") or "",
                    t.get("orden_nro")
                )
                if data:
                    cur.execute("""
                        UPDATE cambios SET orden_envio_tn_id=%s, orden_envio_tn_number=%s WHERE id=%s;
                    """, (data.get("id"), str(data.get("number") or ""), cambio_id))
                    cur.execute("""
                        INSERT INTO cambios_historial (cambio_id, token_id, accion, datos)
                        VALUES (%s,%s,'orden_envio_creada',%s);
                    """, (cambio_id, c["token_id"], json.dumps({"orden_id": data.get("id"), "number": data.get("number")})))
                    envio_info = {"creada": True, "number": data.get("number")}
                else:
                    cur.execute("""
                        INSERT INTO cambios_historial (cambio_id, token_id, accion, datos)
                        VALUES (%s,%s,'orden_envio_error',%s);
                    """, (cambio_id, c["token_id"], json.dumps({"error": err})))
                    envio_info = {"creada": False, "error": err}
            conn.commit()
    return {"ok": True, "envio": envio_info}


class RechazarPayload(BaseModel):
    motivo: str
    aprobado_por: str = "admin"

@app.post("/api/cambios/{cambio_id}/rechazar")
def rechazar_cambio(cambio_id: int, payload: RechazarPayload):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM cambios WHERE id=%s;", (cambio_id,))
            c = cur.fetchone()
            if not c:
                raise HTTPException(status_code=404, detail="Cambio no encontrado")
            cur.execute("""
                UPDATE cambios
                SET estado='rechazado', motivo_rechazo=%s, aprobado_por=%s,
                    aprobado_en=NOW(), actualizado_en=NOW()
                WHERE id=%s;
            """, (payload.motivo, payload.aprobado_por, cambio_id))
            cur.execute("""
                INSERT INTO cambios_historial (cambio_id, token_id, accion, datos)
                VALUES (%s,%s,'rechazado',%s);
            """, (cambio_id, c["token_id"], json.dumps({"motivo": payload.motivo, "por": payload.aprobado_por})))
            conn.commit()
    return {"ok": True}


@app.post("/api/cambios/{cambio_id}/marcar-recibido")
def marcar_recibido(cambio_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT token_id, estado FROM cambios WHERE id=%s;", (cambio_id,))
            c = cur.fetchone()
            if not c:
                raise HTTPException(status_code=404, detail="Cambio no encontrado")
            if c["estado"] != "pendiente_recepcion":
                raise HTTPException(status_code=409, detail=f"Estado actual: {c['estado']}")
            cur.execute("UPDATE cambios SET estado='pendiente_aprobacion', actualizado_en=NOW() WHERE id=%s;", (cambio_id,))
            cur.execute("INSERT INTO cambios_historial (cambio_id, token_id, accion, datos) VALUES (%s,%s,'recibido','{}');", (cambio_id, c["token_id"]))
            conn.commit()
    return {"ok": True}


@app.get("/api/cambios/{cambio_id}/historial")
def historial_cambio(cambio_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT * FROM cambios_historial
                WHERE cambio_id=%s ORDER BY creado_en ASC;
            """, (cambio_id,))
            return cur.fetchall()

@app.get("/cambios/{token}", response_class=HTMLResponse)
def pagina_cambio(token: str):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT token_id, expira_at, finalizado FROM tokens_cambio WHERE token_id=%s;", (token,))
            t = cur.fetchone()
    if not t:
        return HTMLResponse("<div style='font-family:system-ui;padding:3rem;text-align:center'><h2>Link invalido</h2><p style='color:#666;margin-top:.5rem'>El link que abriste no existe.</p></div>", status_code=404)
    if datetime.now() > t["expira_at"]:
        return HTMLResponse("<div style='font-family:system-ui;padding:3rem;text-align:center'><h2>Link expirado</h2><p style='color:#666;margin-top:.5rem'>Este link ya no esta disponible. Contactanos por WhatsApp si necesitas ayuda.</p></div>", status_code=410)
    if t.get("finalizado"):
        return HTMLResponse("<div style='font-family:system-ui;padding:3rem;text-align:center'><h2>Cambio confirmado</h2><p style='color:#666;margin-top:.5rem'>Ya finalizaste tu eleccion. Te vamos a escribir por WhatsApp para coordinar.</p></div>", status_code=200)

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT s.id, s.nombre, s.talle, s.color, s.imagen_url
                FROM stock s
                LEFT JOIN (
                    SELECT remera_elegida_id, COUNT(*) AS reservadas
                    FROM cambios
                    WHERE estado IN ('pendiente_recepcion','pendiente_aprobacion')
                    GROUP BY remera_elegida_id
                ) c ON c.remera_elegida_id = s.id
                WHERE s.cantidad - COALESCE(c.reservadas, 0) > 0
                ORDER BY s.nombre, s.talle;
            """)
            remeras = cur.fetchall()

    remeras_json = json.dumps([{
        "id": r["id"],
        "nombre": str(r["nombre"] or ""),
        "talle": str(r["talle"] or ""),
        "color": str(r["color"] or ""),
        "imagen_url": str(r["imagen_url"] or "")
    } for r in remeras])

    html = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Cambio de remera - Samcro</title>
<style>
:root{
  --black:#0a0a0a; --white:#fafafa;
  --gray-50:#f4f4f4; --gray-100:#e8e8e8; --gray-300:#c8c8c8;
  --gray-400:#9a9a9a; --gray-600:#555;
  --green:#16a34a; --green-light:#f0fdf4;
  --red:#dc2626; --red-light:#fee2e2;
  --radius:12px; --radius-sm:8px;
}
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',system-ui,sans-serif;background:var(--gray-50);color:var(--black);min-height:100vh}
header{background:var(--black);color:var(--white);padding:1rem 1.25rem;position:sticky;top:0;z-index:10;display:flex;align-items:center;justify-content:space-between}
.brand{font-size:.7rem;font-weight:800;letter-spacing:.18em;text-transform:uppercase}
.brand-meta{font-size:.7rem;color:var(--gray-400)}
.screen{display:none;padding:1.25rem 1rem 3rem;max-width:600px;margin:0 auto}
.screen.active{display:block}

/* LOGIN */
.login-wrap{padding-top:2rem}
.login-title{font-size:1.5rem;font-weight:800;letter-spacing:-.02em;margin-bottom:.5rem;line-height:1.2}
.login-sub{font-size:.9rem;color:var(--gray-600);line-height:1.5;margin-bottom:1.75rem}
.field{margin-bottom:1rem}
.field label{display:block;font-size:.75rem;font-weight:600;color:var(--gray-600);margin-bottom:.4rem;letter-spacing:.02em}
.field input{width:100%;padding:.85rem 1rem;border:1.5px solid var(--gray-100);border-radius:var(--radius-sm);font-size:1rem;background:var(--white);transition:border-color .15s}
.field input:focus{outline:none;border-color:var(--black)}
.btn-primary{width:100%;padding:1rem;border-radius:var(--radius-sm);background:var(--black);color:var(--white);border:none;font-size:.95rem;font-weight:700;cursor:pointer;transition:opacity .15s}
.btn-primary:disabled{opacity:.5;cursor:not-allowed}
.btn-secondary{width:100%;padding:.85rem;border-radius:var(--radius-sm);background:var(--white);color:var(--black);border:1.5px solid var(--gray-100);font-size:.9rem;font-weight:600;cursor:pointer;margin-top:.5rem}
.btn-confirm{background:var(--green)}
.error-msg{background:var(--red-light);color:#991b1b;border-radius:var(--radius-sm);padding:.75rem 1rem;font-size:.85rem;margin-bottom:1rem;display:none}
.error-msg.show{display:block}

/* SELECCION */
.cliente-greeting{font-size:.75rem;color:var(--gray-600);margin-bottom:.25rem;letter-spacing:.02em}
.cliente-name{font-size:1.35rem;font-weight:800;letter-spacing:-.02em;line-height:1.2;margin-bottom:.25rem}
.intro{font-size:.85rem;color:var(--gray-600);margin-bottom:1.5rem;line-height:1.5}
.original-card{background:var(--white);border-radius:var(--radius);padding:.85rem;margin-bottom:.85rem;display:flex;gap:.85rem;align-items:flex-start;border:1.5px solid var(--gray-100)}
.original-img{width:72px;height:72px;border-radius:var(--radius-sm);object-fit:cover;background:var(--gray-50);flex-shrink:0}
.original-img-ph{width:72px;height:72px;border-radius:var(--radius-sm);background:var(--gray-100);flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:.65rem;color:var(--gray-400)}
.original-info{flex:1;min-width:0}
.original-label{font-size:.6rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--gray-400);margin-bottom:.25rem}
.original-name{font-size:.9rem;font-weight:700;line-height:1.25;margin-bottom:.2rem}
.original-meta{font-size:.75rem;color:var(--gray-600)}
.swap-arrow{text-align:center;color:var(--gray-300);font-size:1.1rem;margin:-.25rem 0;letter-spacing:.4em}
.replacement-card{background:var(--white);border-radius:var(--radius);padding:.85rem;margin-bottom:1.5rem;border:1.5px solid var(--gray-100);transition:border-color .15s}
.replacement-card.empty{border-style:dashed;background:transparent;cursor:pointer;text-align:center;padding:1.5rem 1rem}
.replacement-card.empty:hover{border-color:var(--black);background:var(--white)}
.btn-elegir{width:100%;padding:1rem;border-radius:var(--radius);background:var(--white);color:var(--black);border:1.5px dashed var(--gray-300);font-size:.9rem;font-weight:600;cursor:pointer;margin-bottom:1.5rem;transition:all .15s}
.btn-elegir:hover{border-color:var(--black);border-style:solid}
.btn-quitar{background:none;border:none;color:var(--red);font-size:.75rem;cursor:pointer;padding:.4rem .6rem;text-decoration:underline}
.no-cambia-pill{display:flex;align-items:center;justify-content:space-between;background:var(--gray-50);color:var(--gray-600);font-size:.8rem;font-weight:500;padding:.75rem 1rem;border-radius:var(--radius);margin-bottom:1.5rem;border:1px solid var(--gray-100)}
.replacement-empty-icon{font-size:1.5rem;margin-bottom:.4rem;color:var(--gray-400)}
.replacement-empty-text{font-size:.85rem;font-weight:600;color:var(--gray-600)}
.replacement-empty-hint{font-size:.75rem;color:var(--gray-400);margin-top:.2rem}
.replacement-content{display:flex;gap:.85rem;align-items:flex-start}
.rep-img{width:72px;height:72px;border-radius:var(--radius-sm);object-fit:cover;background:var(--gray-50);flex-shrink:0}
.rep-info{flex:1;min-width:0}
.rep-label{font-size:.6rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;color:var(--green);margin-bottom:.25rem}
.rep-name{font-size:.9rem;font-weight:700;line-height:1.25;margin-bottom:.2rem}
.rep-meta{font-size:.75rem;color:var(--gray-600)}
.rep-change{background:none;border:none;color:var(--gray-600);font-size:.75rem;cursor:pointer;padding:.25rem 0;text-decoration:underline;margin-top:.4rem}

.bottom-actions{position:sticky;bottom:0;background:linear-gradient(to top,var(--gray-50) 80%,transparent);padding:1rem 0 .25rem;margin-top:1rem}
.bottom-actions .btn-primary{margin-bottom:.5rem}
.guia-link{font-size:.8rem;color:var(--gray-600);background:none;border:none;cursor:pointer;text-decoration:underline;text-underline-offset:3px;display:block;margin:1rem auto 0}

/* PICKER FULLSCREEN */
.picker{display:none;position:fixed;inset:0;background:var(--gray-50);z-index:100;flex-direction:column}
.picker.open{display:flex}
.picker-head{background:var(--black);color:var(--white);padding:1rem 1.25rem;display:flex;align-items:center;gap:.85rem;flex-shrink:0}
.picker-close{background:none;border:none;color:var(--white);font-size:1.4rem;cursor:pointer;padding:0;line-height:1}
.picker-title{font-size:.9rem;font-weight:700;flex:1}
.picker-body{flex:1;overflow-y:auto;padding:1rem;max-width:600px;width:100%;margin:0 auto}
.chips-wrap{display:flex;gap:.35rem;overflow-x:auto;padding-bottom:.5rem;margin-bottom:1rem;scrollbar-width:none}
.chips-wrap::-webkit-scrollbar{display:none}
.chip{border:1.5px solid var(--gray-100);border-radius:20px;padding:.3rem .9rem;font-size:.78rem;font-weight:600;background:var(--white);color:var(--gray-600);cursor:pointer;white-space:nowrap;flex-shrink:0;transition:all .15s}
.chip.sel{border-color:var(--black);background:var(--black);color:var(--white)}
.grid{display:grid;grid-template-columns:repeat(2,1fr);gap:.65rem}
@media(min-width:480px){.grid{grid-template-columns:repeat(3,1fr)}}
.card{background:var(--white);border-radius:var(--radius);overflow:hidden;cursor:pointer;border:2px solid transparent;transition:all .15s}
.card:active{transform:scale(.97)}
.card-img{width:100%;aspect-ratio:1/1;object-fit:contain;background:var(--gray-50);padding:.5rem;display:block}
.card-img-ph{width:100%;aspect-ratio:1/1;background:var(--gray-100);display:flex;align-items:center;justify-content:center;color:var(--gray-400);font-size:.7rem}
.card-body{padding:.5rem .65rem .65rem}
.card-name{font-size:.78rem;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:.15rem}
.card-meta{font-size:.7rem;color:var(--gray-400)}
.card-badge{display:inline-block;background:var(--gray-50);border:1px solid var(--gray-100);border-radius:4px;font-size:.65rem;font-weight:700;padding:.1rem .35rem;margin-top:.3rem;letter-spacing:.03em}
.empty-state{text-align:center;color:var(--gray-400);padding:3rem 1rem;font-size:.85rem}

/* SUCCESS */
.success-wrap{display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:60vh;text-align:center;padding:2rem 1rem}
.success-ring{width:72px;height:72px;border-radius:50%;background:var(--green-light);border:2px solid #bbf7d0;display:flex;align-items:center;justify-content:center;margin:0 auto 1.5rem;color:var(--green);font-size:1.8rem;font-weight:800}
.success-title{font-size:1.5rem;font-weight:800;letter-spacing:-.02em;margin-bottom:.6rem}
.success-body{font-size:.9rem;color:var(--gray-600);line-height:1.7;max-width:320px}

/* MODAL guia talles */
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:200;align-items:flex-end;justify-content:center}
.modal-overlay.open{display:flex}
.modal-sheet{background:var(--white);border-radius:20px 20px 0 0;width:100%;max-width:560px;max-height:90vh;overflow:hidden;display:flex;flex-direction:column}
.modal-handle{width:36px;height:4px;border-radius:2px;background:var(--gray-100);margin:.7rem auto .3rem;flex-shrink:0}
.modal-head{padding:.5rem 1rem .75rem;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid var(--gray-100);flex-shrink:0}
.modal-head h3{font-size:.9rem;font-weight:700}
.modal-close{background:none;border:none;font-size:1.2rem;cursor:pointer;color:var(--gray-600);padding:.2rem}
.modal-sheet iframe{flex:1;border:none;min-height:75vh}
</style>
</head>
<body>

<header>
  <span class="brand">Samcro Remeras</span>
  <span class="brand-meta" id="header-orden"></span>
</header>

<!-- LOGIN -->
<div class="screen active" id="screen-login">
  <div class="login-wrap">
    <h1 class="login-title">Bienvenido al portal de cambios</h1>
    <p class="login-sub">Ingresa el email con el que hiciste tu compra para ver tu orden y elegir el cambio.</p>
    <div class="error-msg" id="login-error"></div>
    <div class="field">
      <label>Email</label>
      <input type="email" id="email-input" placeholder="tucorreo@ejemplo.com" autocomplete="email" inputmode="email">
    </div>
    <button class="btn-primary" id="login-btn" onclick="login()">Continuar</button>
  </div>
</div>

<!-- SELECCION -->
<div class="screen" id="screen-seleccion">
  <p class="cliente-greeting">Hola</p>
  <h1 class="cliente-name" id="cliente-name"></h1>
  <p class="intro">Para cada remera de tu orden, elegi cual queres recibir a cambio.</p>
  <div id="originales-wrap"></div>
  <div class="bottom-actions">
    <button class="btn-primary btn-confirm" id="finalizar-btn" onclick="finalizar()" disabled>Confirmar mis elecciones</button>
    <button class="btn-secondary" onclick="guardarYSalir()">Guardar y volver luego</button>
    <button class="guia-link" onclick="abrirGuia()">Ver guia de talles</button>
  </div>
</div>

<!-- PICKER FULLSCREEN -->
<div class="picker" id="picker">
  <div class="picker-head">
    <button class="picker-close" onclick="cerrarPicker()">&#10005;</button>
    <span class="picker-title">Elegi la remera nueva</span>
    <button class="guia-link" style="color:#fff;margin:0;font-size:.75rem" onclick="abrirGuia()">Talles</button>
  </div>
  <div class="picker-body">
    <div class="chips-wrap" id="picker-chips"></div>
    <div class="grid" id="picker-grid"></div>
  </div>
</div>

<!-- SUCCESS -->
<div class="screen" id="screen-success">
  <div class="success-wrap">
    <div class="success-ring">&#10003;</div>
    <h2 class="success-title">Recibimos tu eleccion</h2>
    <p class="success-body">Te vamos a escribir por WhatsApp cuando recibamos tu prenda original para coordinar el envio del cambio.</p>
  </div>
</div>

<!-- MODAL GUIA -->
<div class="modal-overlay" id="modal-guia" onclick="cerrarGuia()">
  <div class="modal-sheet" onclick="event.stopPropagation()">
    <div class="modal-handle"></div>
    <div class="modal-head">
      <h3>Guia de talles</h3>
      <button class="modal-close" onclick="cerrarGuia()">&#10005;</button>
    </div>
    <iframe src="https://www.samcroremeras.com.ar/guia-de-talles/" title="Guia de talles"></iframe>
  </div>
</div>

<script>
var TOKEN = '""" + str(token) + """';
var STOCK = """ + remeras_json + """;
var EMAIL = '';
var NOMBRE = '';
var ORDEN_NRO = '';
var PRODUCTOS_ORIGINALES = [];
var SELECCIONES = {};
var MARCADAS = {};
var pickerIndex = -1;
var talleFiltro = '';

function showScreen(id) {
  document.querySelectorAll('.screen').forEach(function(s){ s.classList.remove('active'); });
  document.getElementById(id).classList.add('active');
  window.scrollTo(0,0);
}

function abrirGuia(){ document.getElementById('modal-guia').classList.add('open'); }
function cerrarGuia(){ document.getElementById('modal-guia').classList.remove('open'); }

function login() {
  var email = document.getElementById('email-input').value.trim().toLowerCase();
  var err = document.getElementById('login-error');
  var btn = document.getElementById('login-btn');
  err.classList.remove('show');
  if (!email || email.indexOf('@') < 0) {
    err.textContent = 'Ingresa un email valido';
    err.classList.add('show');
    return;
  }
  btn.disabled = true;
  btn.textContent = 'Validando...';
  fetch('/api/validar-acceso', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({token: TOKEN, email: email})
  }).then(function(r){ return r.json().then(function(d){ return {ok: r.ok, status: r.status, data: d}; }); })
    .then(function(res){
      btn.disabled = false;
      btn.textContent = 'Continuar';
      if (!res.ok) {
        err.textContent = res.data.detail || 'Error al validar';
        err.classList.add('show');
        return;
      }
      EMAIL = email;
      NOMBRE = res.data.cliente_nombre || '';
      ORDEN_NRO = res.data.orden_nro || '';
      PRODUCTOS_ORIGINALES = res.data.productos_originales || [];
      var prevs = res.data.selecciones_previas || [];
      prevs.forEach(function(c){
        var po = c.producto_original || {};
        for (var i = 0; i < PRODUCTOS_ORIGINALES.length; i++) {
          var o = PRODUCTOS_ORIGINALES[i];
          if (o.id === po.id || (o.nombre === po.nombre && o.talle === po.talle)) {
            SELECCIONES[i] = {
              id: c.remera_elegida_id,
              nombre: c.remera_elegida_nombre,
              talle: c.remera_elegida_talle,
              color: c.remera_elegida_color,
              imagen_url: c.remera_elegida_imagen
            };
            MARCADAS[i] = true;
            break;
          }
        }
      });
      try { localStorage.setItem('samcro_email_' + TOKEN, email); } catch(e){}
      mostrarSeleccion();
    })
    .catch(function(){
      btn.disabled = false;
      btn.textContent = 'Continuar';
      err.textContent = 'Error de conexion';
      err.classList.add('show');
    });
}

function mostrarSeleccion() {
  document.getElementById('header-orden').textContent = 'Orden #' + ORDEN_NRO;
  document.getElementById('cliente-name').textContent = NOMBRE || '!Hola!';
  renderOriginales();
  showScreen('screen-seleccion');
}

function renderOriginales() {
  var wrap = document.getElementById('originales-wrap');
  wrap.innerHTML = '';
  PRODUCTOS_ORIGINALES.forEach(function(p, i) {
    var orig = document.createElement('div');
    orig.className = 'original-card';
    if (p.imagen) {
      var im = document.createElement('img');
      im.className = 'original-img';
      im.src = p.imagen;
      im.alt = p.nombre || '';
      im.onerror = function(){
        var ph = document.createElement('div');
        ph.className = 'original-img-ph';
        ph.textContent = 'sin foto';
        im.replaceWith(ph);
      };
      orig.appendChild(im);
    } else {
      var ph = document.createElement('div');
      ph.className = 'original-img-ph';
      ph.textContent = 'sin foto';
      orig.appendChild(ph);
    }
    var info = document.createElement('div');
    info.className = 'original-info';
    info.innerHTML =
      '<div class="original-label">Devolves</div>' +
      '<div class="original-name"></div>' +
      '<div class="original-meta"></div>';
    info.querySelector('.original-name').textContent = p.nombre || '(sin nombre)';
    info.querySelector('.original-meta').textContent = (p.talle ? 'Talle ' + p.talle : '') + (p.cantidad > 1 ? ' \u00b7 x' + p.cantidad : '');
    orig.appendChild(info);
    wrap.appendChild(orig);

    var sel = SELECCIONES[i];
    var marcada = MARCADAS[i];
    if (!marcada) {
      var btn = document.createElement('button');
      btn.className = 'btn-elegir';
      btn.textContent = '+ Cambiar esta remera';
      btn.onclick = function(idx){ return function(){ MARCADAS[idx] = true; renderOriginales(); }; }(i);
      wrap.appendChild(btn);
      return;
    }
    var arrow = document.createElement('div');
    arrow.className = 'swap-arrow';
    arrow.textContent = '\u2193';
    wrap.appendChild(arrow);

    var rep = document.createElement('div');
    if (sel) {
      rep.className = 'replacement-card';
      var content = document.createElement('div');
      content.className = 'replacement-content';
      if (sel.imagen_url) {
        var rim = document.createElement('img');
        rim.className = 'rep-img';
        rim.src = sel.imagen_url;
        rim.alt = sel.nombre || '';
        rim.onerror = function(){
          var ph = document.createElement('div');
          ph.className = 'original-img-ph';
          ph.textContent = 'sin foto';
          rim.replaceWith(ph);
        };
        content.appendChild(rim);
      } else {
        var ph = document.createElement('div');
        ph.className = 'original-img-ph';
        ph.textContent = 'sin foto';
        content.appendChild(ph);
      }
      var ri = document.createElement('div');
      ri.className = 'rep-info';
      ri.innerHTML =
        '<div class="rep-label">Recibis</div>' +
        '<div class="rep-name"></div>' +
        '<div class="rep-meta"></div>' +
        '<button class="rep-change">Cambiar eleccion</button>';
      ri.querySelector('.rep-name').textContent = sel.nombre || '';
      ri.querySelector('.rep-meta').textContent = 'Talle ' + (sel.talle || '-') + (sel.color ? ' \u00b7 ' + sel.color : '');
      ri.querySelector('.rep-change').onclick = function(){ abrirPicker(i); };
      content.appendChild(ri);
      rep.appendChild(content);
    } else {
      rep.className = 'replacement-card empty';
      rep.onclick = function(){ abrirPicker(i); };
      rep.innerHTML =
        '<div class="replacement-empty-icon">+</div>' +
        '<div class="replacement-empty-text">Elegi tu remera nueva</div>' +
        '<div class="replacement-empty-hint">Ver opciones disponibles</div>';
    }
    wrap.appendChild(rep);
    var quitarWrap = document.createElement('div');
    quitarWrap.style.textAlign = 'center';
    quitarWrap.style.marginBottom = '1.5rem';
    var quitar = document.createElement('button');
    quitar.className = 'btn-quitar';
    quitar.textContent = 'No cambiar esta';
    quitar.onclick = function(idx){ return function(e){ e.stopPropagation(); delete MARCADAS[idx]; delete SELECCIONES[idx]; renderOriginales(); }; }(i);
    quitarWrap.appendChild(quitar);
    wrap.appendChild(quitarWrap);
  });
  actualizarBotonFinalizar();
}

function actualizarBotonFinalizar() {
  var hayMarcada = false, faltaSel = false;
  for (var k in MARCADAS) {
    if (MARCADAS[k]) { hayMarcada = true; if (!SELECCIONES[k]) faltaSel = true; }
  }
  document.getElementById('finalizar-btn').disabled = !(hayMarcada && !faltaSel);
}

function abrirPicker(index) {
  pickerIndex = index;
  talleFiltro = '';
  var ts = {};
  STOCK.forEach(function(r){ if (r.talle) ts[r.talle] = true; });
  var talles = Object.keys(ts).sort();
  var chips = document.getElementById('picker-chips');
  chips.innerHTML = '';
  function addChip(label, valor) {
    var b = document.createElement('button');
    b.className = 'chip' + (valor === talleFiltro ? ' sel' : '');
    b.textContent = label;
    b.onclick = function(){
      talleFiltro = valor;
      document.querySelectorAll('#picker-chips .chip').forEach(function(c){ c.classList.remove('sel'); });
      b.classList.add('sel');
      renderPickerGrid();
    };
    chips.appendChild(b);
  }
  addChip('Todos', '');
  talles.forEach(function(t){ addChip(t, t); });
  renderPickerGrid();
  document.getElementById('picker').classList.add('open');
}

function renderPickerGrid() {
  var filtradas = talleFiltro ? STOCK.filter(function(r){ return r.talle === talleFiltro; }) : STOCK;
  var g = document.getElementById('picker-grid');
  g.innerHTML = '';
  if (!filtradas.length) {
    g.innerHTML = '<p class="empty-state" style="grid-column:1/-1">No hay remeras disponibles</p>';
    return;
  }
  filtradas.forEach(function(r){
    var card = document.createElement('div');
    card.className = 'card';
    card.onclick = function(){ elegirRemera(r); };
    if (r.imagen_url) {
      var im = document.createElement('img');
      im.className = 'card-img';
      im.src = r.imagen_url;
      im.alt = r.nombre || '';
      im.onerror = function(){
        var ph = document.createElement('div');
        ph.className = 'card-img-ph';
        ph.textContent = 'sin foto';
        im.replaceWith(ph);
      };
      card.appendChild(im);
    } else {
      var ph = document.createElement('div');
      ph.className = 'card-img-ph';
      ph.textContent = 'sin foto';
      card.appendChild(ph);
    }
    var body = document.createElement('div');
    body.className = 'card-body';
    var n = document.createElement('div'); n.className = 'card-name'; n.textContent = r.nombre || '';
    var m = document.createElement('div'); m.className = 'card-meta'; m.textContent = r.color || '';
    var b = document.createElement('span'); b.className = 'card-badge'; b.textContent = r.talle || '';
    body.appendChild(n); body.appendChild(m); body.appendChild(b);
    card.appendChild(body);
    g.appendChild(card);
  });
}

function elegirRemera(r) {
  if (pickerIndex >= 0) {
    SELECCIONES[pickerIndex] = r;
    MARCADAS[pickerIndex] = true;
    cerrarPicker();
    renderOriginales();
  }
}

function cerrarPicker() {
  document.getElementById('picker').classList.remove('open');
  pickerIndex = -1;
}

function _payloadSelecciones() {
  var arr = [];
  Object.keys(SELECCIONES).forEach(function(k){
    arr.push({ producto_original_index: parseInt(k), remera_id: SELECCIONES[k].id });
  });
  return arr;
}

function guardarYSalir() {
  var sel = _payloadSelecciones();
  if (!sel.length) {
    alert('Todavia no elegiste ninguna remera');
    return;
  }
  fetch('/api/cambios/seleccionar', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({token: TOKEN, email: EMAIL, selecciones: sel, finalizar: false})
  }).then(function(r){ return r.json(); })
    .then(function(){ alert('Guardamos tu progreso. Podes volver al link cuando quieras.'); });
}

function finalizar() {
  var sel = _payloadSelecciones();
  if (sel.length === 0) {
    alert('Marca al menos una remera para cambiar');
    return;
  }
  if (!confirm('Vas a cambiar ' + sel.length + ' remera(s). Una vez confirmado no podes modificar tu eleccion. Continuar?')) return;
  var btn = document.getElementById('finalizar-btn');
  btn.disabled = true;
  btn.textContent = 'Confirmando...';
  fetch('/api/cambios/seleccionar', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({token: TOKEN, email: EMAIL, selecciones: sel, finalizar: true})
  }).then(function(r){ return r.json().then(function(d){ return {ok: r.ok, data: d}; }); })
    .then(function(res){
      if (res.ok) { showScreen('screen-success'); }
      else {
        btn.disabled = false;
        btn.textContent = 'Confirmar mis elecciones';
        alert('Error: ' + (res.data.detail || 'no se pudo finalizar'));
      }
    })
    .catch(function(){
      btn.disabled = false;
      btn.textContent = 'Confirmar mis elecciones';
      alert('Error de conexion');
    });
}

window.addEventListener('keydown', function(e){
  if (e.key === 'Enter' && document.getElementById('screen-login').classList.contains('active')) {
    login();
  }
});

(function autoLogin() {
  try {
    var saved = localStorage.getItem('samcro_email_' + TOKEN);
    if (saved) document.getElementById('email-input').value = saved;
  } catch(e){}
})();
</script>
</body>
</html>"""
    return html


@app.get("/api/buscar-productos")
def buscar_productos(q: str = ""):
    if not q or len(q) < 2:
        return []
    res = requests.get(
        f"https://api.tiendanube.com/v1/{TN_STORE_ID}/products",
        headers={
            "Authentication": f"bearer {TN_ACCESS_TOKEN}",
            "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
        },
        params={"q": q, "per_page": 10}
    )
    if res.status_code != 200:
        return []
    productos = res.json()
    resultado = []
    for p in productos:
        nombre = p.get("name", {}).get("es", "") or ""
        imagen = ""
        if p.get("images"):
            imagen = p["images"][0].get("src", "")
        link = p.get("canonical_url", "") or p.get("permalink", "")
        resultado.append({"nombre": nombre, "imagen": imagen, "link": link})
    return resultado

@app.post("/api/actualizar-imagenes")
def actualizar_imagenes():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id, nombre FROM stock WHERE imagen_url IS NULL OR imagen_url = '';")
            remeras = cur.fetchall()
    
    actualizadas = 0
    for r in remeras:
        try:
            res = requests.get(
                f"https://api.tiendanube.com/v1/{TN_STORE_ID}/products",
                headers={
                    "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                    "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
                },
                params={"q": r["nombre"], "per_page": 10, "category_id": 1031807}
            )
            if res.status_code != 200:
                continue
            productos = res.json()
            if not productos:
                continue
            imagen = ""
            link = ""
            for p in productos:
                nombre_tn = p.get("name", {}).get("es", "") or ""
                if nombre_tn.lower().strip() == r["nombre"].lower().strip():
                    if p.get("images"):
                        imagen = p["images"][0].get("src", "")
                    link = p.get("canonical_url", "") or p.get("permalink", "")
                    break
            if imagen or link:
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE stock SET imagen_url=%s, link_tienda=%s WHERE id=%s;",
                            (imagen, link, r["id"])
                        )
                        conn.commit()
                actualizadas += 1
        except:
            continue
    
    return {"ok": True, "actualizadas": actualizadas}
    
@app.get("/api/categorias")
def get_categorias():
    res = requests.get(
        f"https://api.tiendanube.com/v1/{TN_STORE_ID}/categories",
        headers={
            "Authentication": f"bearer {TN_ACCESS_TOKEN}",
            "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
        }
    )
    return res.json()
    
PANEL_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Samcro - Panel de Stock</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#fafaf7;--surface:#ffffff;--surface-2:#f2f2ee;--line:#e5e5e0;--line-2:#c8c8c0;
  --ink:#0e0e0e;--ink-dim:#5a5a55;--ink-mute:#9a9a92;
  --accent:#0e0e0e;--accent-ink:#c8ff2c;--accent-2:#7cc628;--warn:#e85d20;--ok:#2d8f2d;--err:#d83232;
  --mono:'JetBrains Mono','SF Mono','Consolas',monospace;
  --sans:'Inter','Helvetica Neue',system-ui,sans-serif;
}
html,body{height:100%}
body{font-family:var(--sans);background:var(--bg);color:var(--ink);font-size:14px;-webkit-font-smoothing:antialiased}
header{background:var(--bg);border-bottom:1px solid var(--line);padding:0;display:grid;grid-template-columns:auto 1fr auto;align-items:stretch}
header h1{font-family:var(--mono);font-size:.78rem;font-weight:600;letter-spacing:.18em;padding:1.1rem 1.5rem;border-right:1px solid var(--line);display:flex;align-items:center;gap:.5rem}
header h1::before{content:"";width:8px;height:8px;background:var(--accent-2);display:inline-block}
.nav{display:flex;align-items:center}
.nav a{font-family:var(--mono);font-size:.7rem;letter-spacing:.15em;color:var(--ink-dim);padding:0 1.25rem;height:100%;display:flex;align-items:center;text-decoration:none;border-right:1px solid var(--line);text-transform:uppercase;transition:color .15s}
.nav a:hover{color:var(--ink)}
.nav a.current{color:var(--ink);background:var(--surface-2);box-shadow:inset 0 -2px 0 var(--accent-2)}
.nav .bdg{background:var(--warn);color:#fff;font-size:.6rem;padding:1px 6px;margin-left:.5rem;font-weight:700;border-radius:0}
.actions{display:flex;align-items:center;padding-right:1rem;gap:.5rem}
.tb-group{display:flex;gap:.25rem}
.tb-sep{width:1px;height:22px;background:var(--line);margin:0 .35rem}
.btn{font-family:var(--mono);padding:.5rem .9rem;border:1px solid var(--line-2);background:transparent;color:var(--ink);cursor:pointer;font-size:.7rem;font-weight:500;letter-spacing:.12em;text-transform:uppercase;border-radius:0;transition:all .12s}
.btn:hover{border-color:var(--ink);background:var(--surface-2)}
.btn-white{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.btn-white:hover{background:var(--accent);color:var(--accent-ink);border-color:var(--accent)}
.btn-green{background:var(--ink);color:var(--accent-ink);border-color:var(--ink);font-weight:700}
.btn-green:hover{background:#1f1f1f;border-color:#1f1f1f}
.btn-blue{background:transparent;color:var(--ink);border-color:var(--line-2)}
.btn-danger{background:transparent;color:var(--err);border-color:var(--err)}
.btn-danger:hover{background:var(--err);color:#fff;border-color:var(--err)}
main{padding:0}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:0;border-bottom:1px solid var(--line)}
.stat{padding:1.5rem 1.75rem;border-right:1px solid var(--line);position:relative}
.stat:last-child{border-right:none}
.stat p{font-family:var(--mono);font-size:.65rem;color:var(--ink-mute);letter-spacing:.18em;text-transform:uppercase;margin-bottom:.5rem}
.stat h2{font-family:var(--mono);font-size:2.4rem;font-weight:500;letter-spacing:-.02em;line-height:1;color:var(--ink)}
.stat::after{content:"";position:absolute;top:1.5rem;right:1.75rem;width:4px;height:4px;background:var(--accent-2);border-radius:50%}
.toolbar{padding:1rem 1.75rem;border-bottom:1px solid var(--line);display:flex;justify-content:space-between;align-items:center;font-family:var(--mono);font-size:.7rem;color:var(--ink-mute);letter-spacing:.12em;text-transform:uppercase}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:0;padding:0}
.card{background:var(--surface);border-right:1px solid var(--line);border-bottom:1px solid var(--line);overflow:hidden;display:flex;flex-direction:column;transition:background .12s}
.card:hover{background:var(--surface-2)}
.card .img-wrap{position:relative;aspect-ratio:1;background:var(--bg);overflow:hidden}
.card img{width:100%;height:100%;object-fit:cover;display:block;filter:contrast(1.05)}
.card .qty-tag{position:absolute;top:.5rem;left:.5rem;font-family:var(--mono);font-size:.7rem;background:var(--ink);color:#fff;padding:.25rem .5rem;letter-spacing:.1em;border:1px solid var(--ink)}
.card .qty-tag.zero{background:var(--err);border-color:var(--err);color:#fff}
.card-body{padding:.85rem;flex:1;display:flex;flex-direction:column}
.card-body h3{font-size:.85rem;font-weight:500;margin-bottom:.5rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:var(--ink)}
.badges{display:flex;gap:.35rem;margin-bottom:.6rem;flex-wrap:wrap}
.badge{font-family:var(--mono);font-size:.62rem;padding:.18rem .5rem;font-weight:500;letter-spacing:.1em;text-transform:uppercase;border:1px solid var(--line-2);color:var(--ink-dim)}
.bt{color:var(--ink);border-color:var(--ink)}
.bc{color:var(--ink-dim)}
.bs{display:none}
.card-body p{font-size:.72rem;color:var(--ink-mute);margin-bottom:.75rem;font-family:var(--mono);text-transform:uppercase;letter-spacing:.08em}
.card-actions{display:flex;gap:0;margin-top:auto;border-top:1px solid var(--line)}
.card-actions button{flex:1;padding:.6rem .35rem;border:none;background:transparent;color:var(--ink-dim);cursor:pointer;font-size:.62rem;font-family:var(--mono);letter-spacing:.1em;text-transform:uppercase;border-right:1px solid var(--line);transition:all .12s}
.card-actions button:last-child{border-right:none}
.card-actions button:hover{background:var(--bg);color:var(--ink)}
.card-actions button.act-link:hover{color:var(--accent-2)}
.card-actions button.act-del:hover{color:var(--err)}
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.85);z-index:100;align-items:center;justify-content:center;backdrop-filter:blur(4px)}
.modal-bg.open{display:flex}
.modal{background:var(--surface);border:1px solid var(--line-2);padding:1.75rem;width:100%;max-width:460px;max-height:90vh;overflow-y:auto;border-radius:0;position:relative}
.modal::before{content:"";position:absolute;top:0;left:0;right:0;height:2px;background:var(--accent-2)}
.modal h2{font-family:var(--mono);font-size:.78rem;font-weight:600;letter-spacing:.18em;text-transform:uppercase;margin-bottom:1.25rem;color:var(--ink)}
.field{margin-bottom:.85rem}
.field label{display:block;font-family:var(--mono);font-size:.62rem;color:var(--ink-mute);margin-bottom:.35rem;letter-spacing:.15em;text-transform:uppercase}
.field input,.field select,.modal textarea{width:100%;padding:.65rem .75rem;border:1px solid var(--line-2);border-radius:0;font-size:.85rem;background:var(--bg);color:var(--ink);font-family:var(--sans)}
.field input:focus,.field select:focus,.modal textarea:focus{outline:none;border-color:var(--ink);background:var(--surface)}
.field-row{display:grid;grid-template-columns:1fr 1fr;gap:.5rem}
.modal-actions{display:flex;justify-content:flex-end;gap:.5rem;margin-top:1.5rem;padding-top:1rem;border-top:1px solid var(--line)}
.empty{text-align:center;padding:5rem 2rem;color:var(--ink-mute);grid-column:1/-1;font-family:var(--mono);font-size:.75rem;letter-spacing:.15em;text-transform:uppercase}
.sg-item{display:flex;align-items:center;gap:10px;padding:.6rem .75rem;cursor:pointer;border-bottom:1px solid var(--line);background:var(--bg)}
.sg-item:hover{background:var(--surface-2)}
.sg-item span{color:var(--ink);font-size:.82rem}
#sg{border:1px solid var(--line-2)!important;background:var(--bg)!important;border-radius:0!important}
.token-box{background:var(--surface-2);border:1px solid var(--ink);padding:1rem;margin-top:1rem;position:relative}
.token-box::before{content:"// LINK GENERADO";position:absolute;top:-.5rem;left:.75rem;background:var(--surface);padding:0 .5rem;font-family:var(--mono);font-size:.6rem;color:var(--ink);letter-spacing:.15em}
.token-box p{font-size:.75rem;color:var(--ink-dim);margin-bottom:.4rem;font-family:var(--mono)}
.token-box a{color:var(--ink);font-weight:600;word-break:break-all;font-family:var(--mono);font-size:.75rem;text-decoration:underline}
</style>
</head>
<body>
<header>
  <h1>SAMCRO_OPS</h1>
  <nav class="nav">
    <a href="/panel" class="current">Stock</a>
    <a href="/cambios-admin">Cambios <span class="bdg" id="badge-cambios">0</span></a>
  </nav>
  <div class="actions">
    <span class="tb-group">
      <button class="btn" onclick="document.getElementById('fi').click()" title="Importar stock desde Excel">Importar</button>
      <button class="btn" onclick="exportar()" title="Exportar stock a Excel">Exportar</button>
      <button class="btn" onclick="actualizarImagenes()" title="Sincronizar imagenes desde Tienda Nube">Sync imagenes</button>
      <button class="btn btn-danger" onclick="vaciarStock()" title="Borra TODO el stock de la base de datos">Vaciar stock</button>
    </span>
    <span class="tb-sep"></span>
    <button class="btn" onclick="abrirTokenGlobal()" title="Generar link de cambio para un cliente">Generar link de cambio</button>
    <button class="btn btn-green" onclick="abrirModal()">+ Nueva remera</button>
    <input type="file" id="fi" accept=".xlsx" style="display:none" onchange="importar(this)">
  </div>
</header>
<main>
  <div class="stats">
    <div class="stat"><p>SKUs</p><h2 id="st">-</h2></div>
    <div class="stat"><p>Unidades</p><h2 id="su">-</h2></div>
    <div class="stat"><p>Sin stock</p><h2 id="ss">-</h2></div>
  </div>
  <div class="toolbar"><span>// Catalogo</span><span id="tb-count">-</span></div>
  <div class="grid" id="grid"><p class="empty">Cargando...</p></div>
</main>

<div class="modal-bg" id="modal">
  <div class="modal">
    <h2 id="mt">Nueva remera</h2>
    <input type="hidden" id="eid">
    <div class="field">
      <label>Nombre</label>
      <input id="fn" placeholder="Escribi para buscar..." autocomplete="off" oninput="buscar(this.value)">
      <div id="sg" style="border:1px solid #ddd;border-radius:6px;margin-top:4px;display:none;max-height:200px;overflow-y:auto;background:#fff"></div>
    </div>
    <div class="field-row">
      <div class="field"><label>Categoría</label>
        <select id="fcat">
          <option>Música</option>
          <option>Cine y Series</option>
          <option>Superhéroes</option>
          <option>Videojuegos</option>
          <option>Autos y Motos</option>
          <option>Otros</option>
        </select>
      </div>
      <div class="field"><label>Talle</label>
        <select id="ft">
  <option>XS</option><option>S</option><option>M</option>
  <option>L</option><option>XL</option><option>XXL</option>
  <option>XXXL</option><option>4XL</option><option>5XL</option>
  <option>6XL</option><option>7XL</option>
</select>
      </div>
    </div>
    <div class="field-row">
      <div class="field"><label>Color</label><input id="fc" placeholder="negra"></div>
      <div class="field"><label>Cantidad</label><input id="fq" type="number" min="0" value="1"></div>
    </div>
    <div class="field"><label>URL imagen</label><input id="fi2" placeholder="https://..."></div>
    <div class="field"><label>Link tienda</label><input id="fl" placeholder="https://samcroremeras.com.ar/..."></div>
    <div class="modal-actions">
      <button class="btn" onclick="cerrar()">Cancelar</button>
      <button class="btn btn-green" onclick="guardar()">Guardar</button>
    </div>
  </div>
</div>

<div class="modal-bg" id="mtoken">
  <div class="modal">
    <h2>Generar link de cambio</h2>
    <div class="field"><label>Número de orden de Tienda Nube</label><input id="torden" placeholder="10042"></div>
    <p style="font-size:.7rem;color:var(--ink-mute);margin-top:-.4rem;margin-bottom:.85rem;font-family:var(--mono);letter-spacing:.08em">// Buscamos la orden y sus productos en TN automáticamente</p>
    <div class="field" style="background:var(--surface-2);padding:.85rem;border:1px solid var(--line)">
      <p style="font-size:.72rem;color:var(--ink-dim);margin-bottom:.55rem;font-family:var(--mono);letter-spacing:.08em;text-transform:uppercase">// Por defecto: el cliente paga el envío del cambio</p>
      <label style="display:flex;align-items:center;gap:.5rem;cursor:pointer;margin:0;color:var(--ink);text-transform:none;letter-spacing:0;font-family:var(--sans);font-size:.85rem;font-weight:500">
        <input type="checkbox" id="tgratis" style="width:auto;margin:0">
        Cambio sin costo (nosotros cubrimos el envío)
      </label>
      <p style="font-size:.7rem;color:var(--ink-mute);margin-top:.4rem;margin-left:1.5rem">Marcá esta opción solo si el cambio es por cortesía. Si queda sin marcar, al aprobar se genera una orden en TN con "ENVIO CORREO ARGENTINO" para que el cliente abone.</p>
    </div>
    <div class="modal-actions">
      <button class="btn" onclick="document.getElementById('mtoken').classList.remove('open')">Cerrar</button>
      <button class="btn btn-green" id="tbtn" onclick="genToken()">Generar link</button>
    </div>
    <div class="token-box" id="tresult" style="display:none">
      <p style="font-weight:600;margin-bottom:.5rem">Link generado (expira en 5 dias)</p>
      <p style="font-size:.75rem;margin-bottom:.5rem"><strong>Cliente:</strong> <span id="tcli"></span></p>
      <p style="font-size:.75rem;margin-bottom:.5rem"><strong>Email:</strong> <span id="temail"></span></p>
      <p style="font-size:.75rem;margin-bottom:.5rem"><strong>Productos en la orden:</strong> <span id="tprods"></span></p>
      <a id="tlink" href="#" target="_blank" style="display:block;margin-top:.5rem;word-break:break-all"></a>
    </div>
    <div id="terror" style="display:none;background:#fee2e2;color:#991b1b;border-radius:8px;padding:.75rem;margin-top:1rem;font-size:.8rem"></div>
  </div>
</div>
   
<script>
var remeras = [];
var sugs = [];

function esc(s) {
  return (s === null || s === undefined) ? '' : String(s).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/'/g,'&#39;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}

function buscar(q) {
  var box = document.getElementById('sg');
  if (q.length < 2) { box.style.display = 'none'; return; }
  fetch('/api/buscar-productos?q=' + encodeURIComponent(q))
    .then(function(r){ return r.json(); })
    .then(function(items){
      sugs = items;
      if (!items.length) { box.style.display = 'none'; return; }
      box.style.display = 'block';
      var html = '';
      for (var i = 0; i < items.length; i++) {
        html += '<div class="sg-item" onclick="elegir(' + i + ')">';
       html += '<img src="' + esc(items[i].imagen) + '" style="width:40px;height:40px;object-fit:cover;border-radius:4px" onerror="this.style.display=&quot;none&quot;">';
        html += '<span style="font-size:.85rem">' + esc(items[i].nombre) + '</span>';
        html += '</div>';
      }
      box.innerHTML = html;
    });
}

function elegir(i) {
  var p = sugs[i];
  document.getElementById('fn').value = p.nombre;
  document.getElementById('fi2').value = p.imagen;
  document.getElementById('fl').value = p.link;
  document.getElementById('sg').style.display = 'none';
}

function cargar() {
  fetch('/api/stock')
    .then(function(r){ return r.json(); })
    .then(function(data){
      remeras = data;
      renderizar();
    });
  fetch('/api/cambios/pendientes')
    .then(function(r){ return r.json(); })
    .then(function(data){ document.getElementById('badge-cambios').textContent = (data && data.length) || 0; })
    .catch(function(){});
}

function renderizar() {
  var grid = document.getElementById('grid');
  var total = remeras.length;
  var unidades = 0;
  var sin = 0;
  for (var i = 0; i < remeras.length; i++) {
    unidades += remeras[i].cantidad || 0;
    if (!remeras[i].cantidad) sin++;
  }
  document.getElementById('st').textContent = total;
  document.getElementById('su').textContent = unidades;
  document.getElementById('ss').textContent = sin;
  var tbc = document.getElementById('tb-count'); if (tbc) tbc.textContent = total + ' SKU / ' + unidades + ' UNID';
  if (!total) { grid.innerHTML = '<p class="empty">// No hay remeras en stock</p>'; return; }
  var html = '';
  for (var i = 0; i < remeras.length; i++) {
    var r = remeras[i];
    var qty = r.cantidad || 0;
    html += '<div class="card">';
    html += '<div class="img-wrap">';
    html += '<span class="qty-tag' + (qty===0?' zero':'') + '">x' + qty + '</span>';
    html += '<img src="' + esc(r.imagen_url) + '" onerror="this.style.display=&quot;none&quot;" alt="">';
    html += '</div>';
    html += '<div class="card-body">';
    html += '<h3 title="' + esc(r.nombre) + '">' + esc(r.nombre) + '</h3>';
    html += '<div class="badges">';
    html += '<span class="badge bt">' + esc(r.talle) + '</span>';
    html += '<span class="badge bc">' + esc(r.categoria) + '</span>';
    html += '</div>';
    html += '<p>' + esc(r.color || '\u2014') + '</p>';
    html += '<div class="card-actions">';
    html += '<button onclick="editar(' + r.id + ')">Editar</button>';
    html += '<button class="act-del" onclick="eliminar(' + r.id + ')">Eliminar</button>';
    html += '</div></div></div>';
  }
  grid.innerHTML = html;
}

function abrirModal() {
  document.getElementById('mt').textContent = 'Nueva remera';
  document.getElementById('eid').value = '';
  document.getElementById('fn').value = '';
  document.getElementById('fc').value = '';
  document.getElementById('fi2').value = '';
  document.getElementById('fl').value = '';
  document.getElementById('fq').value = 1;
  document.getElementById('sg').style.display = 'none';
  document.getElementById('modal').classList.add('open');
}

function cerrar() { document.getElementById('modal').classList.remove('open'); }

function editar(id) {
  var r = null;
  for (var i = 0; i < remeras.length; i++) { if (remeras[i].id === id) { r = remeras[i]; break; } }
  if (!r) return;
  document.getElementById('mt').textContent = 'Editar remera';
  document.getElementById('eid').value = id;
  document.getElementById('fn').value = r.nombre || '';
  document.getElementById('fcat').value = r.categoria || 'Musica';
  document.getElementById('ft').value = r.talle || 'M';
  document.getElementById('fc').value = r.color || '';
  document.getElementById('fq').value = r.cantidad || 0;
  document.getElementById('fi2').value = r.imagen_url || '';
  document.getElementById('fl').value = r.link_tienda || '';
  document.getElementById('modal').classList.add('open');
}

function guardar() {
  var id = document.getElementById('eid').value;
  var data = {
    nombre: document.getElementById('fn').value,
    categoria: document.getElementById('fcat').value,
    talle: document.getElementById('ft').value,
    color: document.getElementById('fc').value,
    cantidad: parseInt(document.getElementById('fq').value) || 0,
    imagen_url: document.getElementById('fi2').value,
    link_tienda: document.getElementById('fl').value
  };
  var url = id ? '/api/stock/' + id : '/api/stock';
  var method = id ? 'PUT' : 'POST';
  fetch(url, {method: method, headers: {'Content-Type': 'application/json'}, body: JSON.stringify(data)})
    .then(function(){ cerrar(); cargar(); });
}

function eliminar(id) {
  if (!confirm('Eliminar esta remera?')) return;
  fetch('/api/stock/' + id, {method: 'DELETE'}).then(function(){ cargar(); });
}

function importar(input) {
  var fd = new FormData();
  fd.append('file', input.files[0]);
  fetch('/api/importar-excel', {method: 'POST', body: fd})
    .then(function(r){ return r.json(); })
    .then(function(data){
      var msg = 'Importadas: ' + data.importadas + ' remeras';
      if (data.hojas) {
        msg += '\\n\\nPor hoja:';
        Object.keys(data.hojas).forEach(function(h){ msg += '\\n  \u2022 ' + h + ': ' + data.hojas[h]; });
      }
      alert(msg); input.value = ''; cargar();
    });
}

function exportar() { window.location.href = '/api/exportar-excel'; }

function vaciarStock(){
  if (!confirm('\u26A0\uFE0F PASO 1/3\\n\\nEsto va a BORRAR TODO el stock de la base de datos.\\nNo se puede deshacer.\\n\\n\u00BFSeguir?')) return;
  if (!confirm('\u26A0\uFE0F PASO 2/3\\n\\nREALMENTE estas seguro? Se pierden TODAS las remeras cargadas.\\n\\nRecomendado: antes exportar a Excel.\\n\\n\u00BFConfirmar?')) return;
  var t = prompt('PASO 3/3\\n\\nEscribi exactamente:\\nBORRAR TODO EL STOCK\\n\\npara confirmar:');
  if (t !== 'BORRAR TODO EL STOCK') { alert('Cancelado. El texto no coincide.'); return; }
  fetch('/api/stock/vaciar?confirmacion=' + encodeURIComponent(t), {method:'POST'})
    .then(function(r){ return r.json().then(function(d){ return {ok:r.ok,d:d}; }); })
    .then(function(x){
      if (!x.ok) { alert('Error: ' + (x.d.detail || '')); return; }
      alert('Stock vaciado. Se borraron ' + x.d.borradas + ' remeras.');
      cargar();
    });
}

function abrirTokenGlobal() {
  document.getElementById('torden').value = '';
  document.getElementById('tgratis').checked = false;
  document.getElementById('tresult').style.display = 'none';
  document.getElementById('terror').style.display = 'none';
  document.getElementById('mtoken').classList.add('open');
  setTimeout(function(){ document.getElementById('torden').focus(); }, 50);
}

function genToken() {
  var orden = document.getElementById('torden').value.trim();
  if (!orden) { alert('Ingresa el numero de orden'); return; }
  var btn = document.getElementById('tbtn');
  btn.disabled = true;
  btn.textContent = 'Buscando orden...';
  document.getElementById('tresult').style.display = 'none';
  document.getElementById('terror').style.display = 'none';
  var pagaEnvio = document.getElementById('tgratis').checked ? 'false' : 'true';
  fetch('/api/tokens?orden_nro=' + encodeURIComponent(orden) + '&cliente_paga_envio=' + pagaEnvio, {method: 'POST'})
    .then(function(r){ return r.json().then(function(d){ return {ok: r.ok, status: r.status, data: d}; }); })
    .then(function(res){
      btn.disabled = false;
      btn.textContent = 'Generar link';
      if (!res.ok) {
        document.getElementById('terror').textContent = 'Error: ' + (res.data.detail || 'no se pudo generar el link');
        document.getElementById('terror').style.display = 'block';
        return;
      }
      var d = res.data;
      document.getElementById('tcli').textContent = d.cliente_nombre || '(sin nombre)';
      document.getElementById('temail').textContent = d.cliente_email || '(sin email)';
      var prods = (d.productos_originales || []).map(function(p){
        return p.nombre + (p.talle ? ' (' + p.talle + ')' : '');
      }).join(', ') || '(sin productos)';
      document.getElementById('tprods').textContent = prods;
      document.getElementById('tlink').textContent = d.link;
      document.getElementById('tlink').href = d.link;
      document.getElementById('tresult').style.display = 'block';
    })
    .catch(function(){
      btn.disabled = false;
      btn.textContent = 'Generar link';
      document.getElementById('terror').textContent = 'Error de conexion';
      document.getElementById('terror').style.display = 'block';
    });
}

function actualizarImagenes() {
  if (!confirm('Esto puede tardar 1-2 minutos. Continuar?')) return;
  fetch('/api/actualizar-imagenes', {method: 'POST'})
    .then(function(r){ return r.json(); })
    .then(function(data){ alert('Actualizadas: ' + data.actualizadas + ' remeras'); cargar(); });
}
cargar();
</script>
</body>
</html>"""
@app.get("/api/productos-remeras")
def listar_productos_remeras():
    todos = []
    page = 1
    while True:
        res = requests.get(
            f"https://api.tiendanube.com/v1/{TN_STORE_ID}/products",
            headers={
                "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                "User-Agent": "Samcro Stock (samcroremeras@gmail.com)"
            },
            params={"category_id": 1031807, "per_page": 200, "page": page}
        )
        if res.status_code != 200:
            break
        data = res.json()
        if not data:
            break
        for p in data:
            todos.append({
                "id": p["id"],
                "nombre": p.get("name", {}).get("es", "") or "",
                "imagen": p["images"][0]["src"] if p.get("images") else ""
            })
        if len(data) < 200:
            break
        page += 1
    return todos
class SubirImagenPayload(BaseModel):
    product_ids: list
    filename: str
    attachment: str

@app.post("/api/subir-tabla-talles")
def subir_tabla_talles(payload: SubirImagenPayload):
    ok = []
    errores = []
    for pid in payload.product_ids:
        res = requests.post(
            f"https://api.tiendanube.com/v1/{TN_STORE_ID}/products/{pid}/images",
            headers={
                "Authentication": f"bearer {TN_ACCESS_TOKEN}",
                "User-Agent": "Samcro Stock (samcroremeras@gmail.com)",
                "Content-Type": "application/json"
            },
            json={"attachment": payload.attachment, "filename": payload.filename}
        )
        if res.status_code in (200, 201):
            ok.append(pid)
        else:
            errores.append({"id": pid, "error": res.text})
    return {"ok": len(ok), "errores": errores}
@app.get("/panel", response_class=HTMLResponse)
def panel():
    return PANEL_HTML


CAMBIOS_ADMIN_HTML = """<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Samcro - Gestion de Cambios</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#fafaf7;--surface:#ffffff;--surface-2:#f2f2ee;--line:#e5e5e0;--line-2:#c8c8c0;
  --ink:#0e0e0e;--ink-dim:#5a5a55;--ink-mute:#9a9a92;
  --accent:#0e0e0e;--accent-ink:#c8ff2c;--accent-2:#7cc628;--warn:#e85d20;--ok:#2d8f2d;--err:#d83232;
  --mono:'JetBrains Mono','SF Mono','Consolas',monospace;
  --sans:'Inter','Helvetica Neue',system-ui,sans-serif;
}
body{font-family:var(--sans);background:var(--bg);color:var(--ink);font-size:14px;-webkit-font-smoothing:antialiased}
header{background:var(--bg);border-bottom:1px solid var(--line);display:grid;grid-template-columns:auto 1fr auto;align-items:stretch}
header h1{font-family:var(--mono);font-size:.78rem;font-weight:600;letter-spacing:.18em;padding:1.1rem 1.5rem;border-right:1px solid var(--line);display:flex;align-items:center;gap:.5rem}
header h1::before{content:"";width:8px;height:8px;background:var(--warn);display:inline-block}
.nav{display:flex;align-items:center}
.nav a{font-family:var(--mono);font-size:.7rem;letter-spacing:.15em;color:var(--ink-dim);padding:0 1.25rem;height:100%;display:flex;align-items:center;text-decoration:none;border-right:1px solid var(--line);text-transform:uppercase;transition:color .15s}
.nav a:hover{color:var(--ink)}
.nav a.current{color:var(--ink);background:var(--surface);box-shadow:inset 0 -2px 0 var(--warn)}
.actions{display:flex;align-items:center;padding-right:1rem}
.btn{font-family:var(--mono);padding:.5rem .9rem;border:1px solid var(--line-2);background:transparent;color:var(--ink);cursor:pointer;font-size:.7rem;font-weight:500;letter-spacing:.12em;text-transform:uppercase;border-radius:0;text-decoration:none;display:inline-block;transition:all .12s}
.btn:hover{border-color:var(--ink);background:var(--surface-2)}
.btn-white{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.btn-white:hover{background:var(--accent);color:var(--accent-ink);border-color:var(--accent)}
.btn-green{background:var(--accent-2);color:#fff;border-color:var(--accent-2);font-weight:700}
.btn-green:hover{background:#5fa61f;border-color:#5fa61f}
.btn-red{background:transparent;color:var(--err);border-color:var(--err)}
.btn-red:hover{background:var(--err);color:var(--bg)}
.btn-blue{background:var(--ink);color:var(--bg);border-color:var(--ink)}
.btn-blue:hover{background:var(--ink-dim);border-color:var(--ink-dim)}
.btn-gray{border-color:var(--line-2);color:var(--ink-dim)}
main{padding:0;max-width:none;margin:0}
.tabs{display:flex;border-bottom:1px solid var(--line);background:var(--bg);overflow-x:auto}
.tab{padding:1.1rem 1.5rem;cursor:pointer;border:none;background:none;font-family:var(--mono);font-size:.7rem;color:var(--ink-mute);font-weight:500;letter-spacing:.15em;text-transform:uppercase;border-right:1px solid var(--line);position:relative;transition:color .15s;white-space:nowrap}
.tab:hover{color:var(--ink-dim)}
.tab.active{color:var(--ink);background:var(--surface)}
.tab.active::after{content:"";position:absolute;left:0;right:0;bottom:0;height:2px;background:var(--warn)}
.tab .count{background:var(--warn);color:#fff;padding:1px 6px;margin-left:.5rem;font-size:.6rem;font-weight:700;letter-spacing:.05em}
.list{display:flex;flex-direction:column}
.cambio{background:var(--surface);border-bottom:1px solid var(--line);transition:background .12s;position:relative}
.cambio-head{padding:1rem 1.75rem;display:grid;grid-template-columns:auto 1fr 1fr auto auto;gap:1.25rem;align-items:center;cursor:pointer;user-select:none}
.cambio-head:hover{background:var(--surface-2)}
.cambio-head .ord{font-family:var(--mono);font-size:.95rem;font-weight:600;color:var(--ink);letter-spacing:.04em}
.cambio-head .ord::before{content:"#";color:var(--ink-mute);margin-right:.1em}
.cambio-head .cli{font-size:.8rem;color:var(--ink-dim);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cambio-head .date{font-family:var(--mono);font-size:.68rem;color:var(--ink-mute);letter-spacing:.08em;text-transform:uppercase;text-align:right}
.cambio-head .chev{font-family:var(--mono);color:var(--ink-mute);transition:transform .15s;font-size:.8rem}
.cambio.open .chev{transform:rotate(90deg)}
.cambio-body{display:none;padding:1rem 1.75rem 1.5rem;border-top:1px solid var(--line);background:var(--bg)}
.cambio.open .cambio-body{display:block}
.cambio-row{display:grid;grid-template-columns:64px 1fr 24px 1fr auto;gap:1.25rem;align-items:center}
.cambio:hover{background:var(--surface-2)}
.cambio::before{content:"";position:absolute;left:0;top:0;bottom:0;width:2px;background:transparent}
.cambio.s-rec::before{background:var(--warn)}
.cambio.s-apr::before{background:var(--accent-2)}
.cambio.s-ok::before{background:var(--ok)}
.cambio.s-no::before{background:var(--err)}
.cambio img{width:64px;height:64px;object-fit:cover;background:var(--bg);border:1px solid var(--line)}
.cambio .info{font-size:.85rem;min-width:0}
.cambio .info strong{display:block;margin-bottom:.25rem;font-weight:500;color:var(--ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cambio .info p{color:var(--ink-mute);font-size:.7rem;font-family:var(--mono);text-transform:uppercase;letter-spacing:.1em}
.cambio .arrow{font-family:var(--mono);font-size:1.25rem;color:var(--warn);text-align:center;font-weight:300}
.cambio .actions{display:flex;flex-direction:column;gap:.4rem;align-items:stretch}
.cambio .actions .btn{font-size:.62rem;padding:.4rem .75rem}
.meta{grid-column:1/-1;border-top:1px solid var(--line);padding-top:.75rem;margin-top:.25rem;font-size:.68rem;color:var(--ink-mute);display:flex;gap:1.5rem;flex-wrap:wrap;font-family:var(--mono);text-transform:uppercase;letter-spacing:.1em}
.meta strong{color:var(--ink-dim);font-weight:500;margin-right:.4rem}
.estado{display:inline-block;padding:.3rem .65rem;font-size:.62rem;font-weight:600;font-family:var(--mono);letter-spacing:.12em;text-transform:uppercase;border:1px solid}
.e-rec{color:var(--warn);border-color:var(--warn)}
.e-apr{color:var(--accent-2);border-color:var(--accent-2)}
.e-ok{color:var(--ok);border-color:var(--ok)}
.e-no{color:var(--err);border-color:var(--err)}
.empty{text-align:center;padding:5rem 2rem;color:var(--ink-mute);font-family:var(--mono);font-size:.75rem;letter-spacing:.15em;text-transform:uppercase;background:var(--surface)}
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.85);z-index:100;align-items:center;justify-content:center;backdrop-filter:blur(4px)}
.modal-bg.open{display:flex}
.modal{background:var(--surface);border:1px solid var(--line-2);padding:1.75rem;width:100%;max-width:460px;border-radius:0;position:relative}
.modal::before{content:"";position:absolute;top:0;left:0;right:0;height:2px;background:var(--err)}
.modal h2{font-family:var(--mono);font-size:.78rem;font-weight:600;letter-spacing:.18em;text-transform:uppercase;margin-bottom:1rem}
.modal p{font-size:.78rem;color:var(--ink-dim);margin-bottom:.85rem}
.modal textarea{width:100%;padding:.65rem .75rem;border:1px solid var(--line-2);border-radius:0;font-size:.85rem;min-height:90px;resize:vertical;font-family:var(--sans);background:var(--bg);color:var(--ink)}
.modal textarea:focus{outline:none;border-color:var(--err)}
.modal-actions{display:flex;justify-content:flex-end;gap:.5rem;margin-top:1.25rem;padding-top:1rem;border-top:1px solid var(--line)}
@media(max-width:768px){.cambio{grid-template-columns:64px 1fr;gap:.75rem}.cambio .arrow,.cambio>.info:nth-of-type(2){grid-column:1/-1}.cambio .arrow{transform:rotate(90deg);text-align:left}.cambio .actions{grid-column:1/-1;flex-direction:row}.cambio .actions .btn{flex:1}}
</style></head><body>
<header>
  <h1>SAMCRO_OPS</h1>
  <nav class="nav">
    <a href="/panel">Stock</a>
    <a href="/cambios-admin" class="current">Cambios</a>
  </nav>
  <div class="actions"></div>
</header>
<main>
  <div class="tabs">
    <button class="tab active" data-estado="" onclick="setTab(this,'')">// Activos <span class="count" id="c-pen">0</span></button>
    <button class="tab" data-estado="pendiente_recepcion" onclick="setTab(this,'pendiente_recepcion')">Esperando recepcion</button>
    <button class="tab" data-estado="pendiente_aprobacion" onclick="setTab(this,'pendiente_aprobacion')">Por aprobar</button>
    <button class="tab" data-estado="aprobado" onclick="setTab(this,'aprobado')">Aprobados</button>
    <button class="tab" data-estado="rechazado" onclick="setTab(this,'rechazado')">Rechazados</button>
  </div>
  <div id="lista" class="list"><p class="empty">// Cargando...</p></div>
</main>

<div class="modal-bg" id="mrech">
  <div class="modal">
    <h2>Rechazar cambio</h2>
    <p style="font-size:.8rem;color:#666;margin-bottom:.75rem">El cliente recibira este motivo. La remera reservada vuelve a estar disponible.</p>
    <textarea id="mot" placeholder="Ej: la remera devuelta llego con manchas / talle incorrecto"></textarea>
    <div class="modal-actions">
      <button class="btn btn-gray" onclick="document.getElementById('mrech').classList.remove('open')">Cancelar</button>
      <button class="btn btn-red" onclick="confirmarRechazo()">Rechazar</button>
    </div>
  </div>
</div>

<script>
var ESTADO = '';
var RECH_ID = null;

function esc(s){ return (s===null||s===undefined)?'':String(s).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/'/g,'&#39;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

function setTab(el, estado){
  document.querySelectorAll('.tab').forEach(function(t){ t.classList.remove('active'); });
  el.classList.add('active');
  ESTADO = estado;
  cargar();
}

function toggleCambio(id, ev){
  if (ev && ev.target && ev.target.closest && ev.target.closest('button')) return;
  var el = document.getElementById('cambio-' + id);
  if (el) el.classList.toggle('open');
}

function fmtFecha(s){ if(!s) return ''; var d = new Date(s); return d.toLocaleDateString('es-AR') + ' ' + d.toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit'}); }

function estadoPill(e){
  var map = {pendiente_recepcion:['e-rec','Esperando recepcion'], pendiente_aprobacion:['e-apr','Por aprobar'], aprobado:['e-ok','Aprobado'], rechazado:['e-no','Rechazado']};
  var v = map[e] || ['e-rec', e];
  return '<span class="estado ' + v[0] + '">' + v[1] + '</span>';
}

function cargar(){
  var url = '/api/cambios/pendientes' + (ESTADO ? '?estado=' + ESTADO : '');
  fetch(url).then(function(r){ return r.json(); }).then(function(data){
    if (!ESTADO) document.getElementById('c-pen').textContent = data.length;
    var c = document.getElementById('lista');
    if (!data.length) { c.innerHTML = '<p class="empty">No hay cambios en este estado.</p>'; return; }
    c.innerHTML = data.map(function(it){
      var po = it.producto_original || {};
      var acciones = '';
      if (it.estado === 'pendiente_recepcion') {
        acciones = '<button class="btn btn-blue" onclick="recibido(' + it.id + ')">Marcar recibido</button>' +
                   '<button class="btn btn-red" onclick="abrirRechazo(' + it.id + ')">Rechazar</button>';
      } else if (it.estado === 'pendiente_aprobacion') {
        acciones = '<button class="btn btn-green" onclick="aprobar(' + it.id + ')">Aprobar</button>' +
                   '<button class="btn btn-red" onclick="abrirRechazo(' + it.id + ')">Rechazar</button>';
      } else {
        acciones = estadoPill(it.estado);
      }
      var cls = {pendiente_recepcion:'s-rec',pendiente_aprobacion:'s-apr',aprobado:'s-ok',rechazado:'s-no'}[it.estado] || '';
      return '<div class="cambio ' + cls + '" id="cambio-' + it.id + '">' +
        '<div class="cambio-head" onclick="toggleCambio(' + it.id + ',event)">' +
          '<span class="ord">' + esc(it.orden_nro||'-') + '</span>' +
          '<span class="cli">' + esc(it.cliente_nombre || it.cliente_email || '-') + '</span>' +
          estadoPill(it.estado) +
          '<span class="date">' + esc(fmtFecha(it.creado_en)) + '</span>' +
          '<span class="chev">\u25B6</span>' +
        '</div>' +
        '<div class="cambio-body">' +
          '<div class="cambio-row">' +
            '<img src="' + esc(po.imagen||'') + '" onerror="this.style.background=\\'#222\\';this.removeAttribute(\\'src\\')">' +
            '<div class="info"><strong>' + esc(po.nombre||'(sin nombre)') + '</strong>' +
              '<p>Devuelve / Talle ' + esc(po.talle||'-') + '</p></div>' +
            '<div class="arrow">\u2192</div>' +
            '<div class="info"><strong>' + esc(it.remera_elegida_nombre||'') + '</strong>' +
              '<p>Recibe / Talle ' + esc(it.remera_elegida_talle||'-') + (it.remera_elegida_color ? ' / ' + esc(it.remera_elegida_color) : '') + '</p></div>' +
            '<div class="actions">' + acciones + '</div>' +
          '</div>' +
          '<div class="meta">' +
            '<span><strong>EMAIL</strong>' + esc(it.cliente_email||'-') + '</span>' +
            '<span style="color:' + (it.cliente_paga_envio ? 'var(--warn)' : 'var(--ink-mute)') + '"><strong>ENVIO</strong>' + (it.cliente_paga_envio ? 'Paga cliente' : 'Sin costo (cubrimos)') + '</span>' +
            (it.orden_envio_tn_number ? '<span style="color:var(--accent-2)"><strong>ENVIO TN</strong>#' + esc(it.orden_envio_tn_number) + '</span>' : '') +
            (it.motivo_rechazo ? '<span style="color:var(--err)"><strong>MOTIVO</strong>' + esc(it.motivo_rechazo) + '</span>' : '') +
          '</div>' +
        '</div>' +
      '</div>';
    }).join('');
  });
}

function recibido(id){
  fetch('/api/cambios/' + id + '/marcar-recibido', {method:'POST'})
    .then(function(r){ return r.json().then(function(d){ return {ok:r.ok,d:d}; }); })
    .then(function(x){ if(!x.ok){ alert(x.d.detail || 'Error'); return; } cargar(); });
}

function aprobar(id){
  if (!confirm('Aprobar este cambio? Se descuenta la remera nueva del stock y se suma la devuelta.')) return;
  fetch('/api/cambios/' + id + '/aprobar', {method:'POST', headers:{'Content-Type':'application/json'}, body:'{}'})
    .then(function(r){ return r.json(); })
    .then(function(){ cargar(); });
}

function abrirRechazo(id){ RECH_ID = id; document.getElementById('mot').value = ''; document.getElementById('mrech').classList.add('open'); }

function confirmarRechazo(){
  var m = document.getElementById('mot').value.trim();
  if (!m) { alert('Escribi un motivo'); return; }
  fetch('/api/cambios/' + RECH_ID + '/rechazar', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({motivo: m})})
    .then(function(r){ return r.json(); })
    .then(function(){ document.getElementById('mrech').classList.remove('open'); cargar(); });
}

cargar();
</script>
</body></html>"""


@app.get("/cambios-admin", response_class=HTMLResponse)
def cambios_admin():
    return CAMBIOS_ADMIN_HTML

TABLA_TALLES_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Subir tabla de talles - Samcro</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:#f5f5f5;color:#111}
header{background:#111;color:#fff;padding:1rem 2rem}
header h1{font-size:1.1rem;font-weight:600}
main{padding:1.5rem 2rem;max-width:900px}
.card{background:#fff;border-radius:8px;border:1px solid #e5e5e5;padding:1.5rem;margin-bottom:1rem}
.btn{padding:.6rem 1.2rem;border-radius:6px;border:none;cursor:pointer;font-size:.9rem;font-weight:500}
.btn-green{background:#16a34a;color:#fff}
.btn-blue{background:#2563eb;color:#fff}
.drop{border:2px dashed #ddd;border-radius:8px;padding:2rem;text-align:center;cursor:pointer;margin-bottom:1rem}
.drop.over{border-color:#2563eb;background:#eff6ff}
.lista{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.75rem;margin-top:1rem}
.prod-item{background:#fff;border:1.5px solid #e5e5e5;border-radius:8px;overflow:hidden;cursor:pointer}
.prod-item.sel{border-color:#111;background:#f9f9f9}
.prod-item img{width:100%;height:130px;object-fit:contain;background:#f5f5f5;padding:4px}
.prod-item .pinfo{padding:.5rem;display:flex;align-items:center;gap:.4rem}
.prod-item input[type=checkbox]{width:16px;height:16px;cursor:pointer}
.prod-item span{font-size:.78rem;font-weight:500;line-height:1.3}
.barra{background:#e5e5e5;border-radius:4px;height:10px;margin-top:.5rem}
.progreso{height:10px;border-radius:4px;background:#16a34a;width:0;transition:width .3s}
.log{font-size:.8rem;color:#666;margin-top:.5rem;max-height:150px;overflow-y:auto}
.acciones{display:flex;gap:.5rem;align-items:center;margin-bottom:.75rem;flex-wrap:wrap}
</style>
</head>
<body>
<header><h1>SAMCRO — Subir tabla de talles</h1></header>
 <div class="card">
    <h2 style="font-size:1rem;margin-bottom:.75rem">4. Eliminar ultima imagen</h2>
    <button class="btn" onclick="eliminarUltima()" id="btn-eliminar" disabled style="background:#fee2e2;color:#dc2626">Eliminar ultima imagen de productos seleccionados</button>
    <div class="barra" style="margin-top:1rem;display:none" id="barra-cont2">
      <div class="progreso" id="progreso2"></div>
    </div>
    <div class="log" id="log2"></div>
  </div>
</main>
  <div class="card">
    <h2 style="font-size:1rem;margin-bottom:1rem">1. Elegi la imagen de la tabla de talles</h2>
    <div class="drop" id="drop" onclick="document.getElementById('fi').click()"
         ondragover="event.preventDefault();this.classList.add('over')"
         ondragleave="this.classList.remove('over')"
         ondrop="onDrop(event)">
      <p style="color:#666;font-size:.9rem">Arrastra o hace click para subir la imagen</p>
      <input type="file" id="fi" accept="image/*" style="display:none" onchange="onFile(this.files[0])">
    </div>
    <img id="prev" style="max-height:180px;max-width:100%;border-radius:6px;display:none;margin:0 auto">
    <p id="fname" style="font-size:.8rem;color:#666;margin-top:.5rem;text-align:center"></p>
  </div>
  <div class="card">
    <h2 style="font-size:1rem;margin-bottom:1rem">2. Selecciona los productos</h2>
    <div class="acciones">
      <button class="btn btn-blue" onclick="cargarProductos()" id="btn-cargar">Cargar productos</button>
      <button class="btn" onclick="toggleTodos(true)" style="background:#f0f0f0">Seleccionar todos</button>
      <button class="btn" onclick="toggleTodos(false)" style="background:#f0f0f0">Deseleccionar todos</button>
      <span id="contador" style="font-size:.85rem;color:#666"></span>
    </div>
    <div class="lista" id="lista"><p style="color:#999;font-size:.85rem">Carga los productos primero.</p></div>
  </div>
  <div class="card">
    <h2 style="font-size:1rem;margin-bottom:.75rem">3. Subir</h2>
    <button class="btn btn-green" onclick="subir()" id="btn-subir" disabled>Subir imagen a productos seleccionados</button>
    <div class="barra" style="margin-top:1rem;display:none" id="barra-cont">
      <div class="progreso" id="progreso"></div>
    </div>
    <div class="log" id="log"></div>
  </div>
</main>
<script>
var imgBase64 = '';
var imgFilename = '';
var productos = [];
var seleccionados = new Set();

function onFile(file) {
  if (!file) return;
  imgFilename = file.name;
  document.getElementById('fname').textContent = file.name;
  var reader = new FileReader();
  reader.onload = function(e) {
    var dataUrl = e.target.result;
    imgBase64 = dataUrl.split(',')[1];
    var prev = document.getElementById('prev');
    prev.src = dataUrl;
    prev.style.display = 'block';
  };
  reader.readAsDataURL(file);
}

function onDrop(e) {
  e.preventDefault();
  document.getElementById('drop').classList.remove('over');
  var f = e.dataTransfer.files[0];
  if (f) onFile(f);
}

function cargarProductos() {
  document.getElementById('lista').innerHTML = '<p style="color:#999;font-size:.85rem">Cargando...</p>';
  document.getElementById('btn-cargar').disabled = true;
  fetch('/api/productos-remeras')
    .then(function(r){ return r.json(); })
    .then(function(data){
      productos = data;
      seleccionados = new Set(data.map(function(p){ return p.id; }));
      renderLista();
      actualizarContador();
      document.getElementById('btn-subir').disabled = false;
      document.getElementById('btn-eliminar').disabled = false;
      document.getElementById('btn-cargar').disabled = false;
    });
}

function renderLista() {
  var html = '';
  productos.forEach(function(p) {
    var sel = seleccionados.has(p.id);
    html += '<div class="prod-item' + (sel ? ' sel' : '') + '" id="pi-' + p.id + '" onclick="toggle(' + p.id + ')">';
    html += '<img src="' + (p.imagen || '') + '" onerror="this.style.display=&quot;none&quot;" alt="">';
    html += '<div class="pinfo">';
    html += '<input type="checkbox" ' + (sel ? 'checked' : '') + ' onclick="event.stopPropagation();toggle(' + p.id + ')">';
    html += '<span>' + esc(p.nombre) + '</span>';
    html += '</div></div>';
  });
  document.getElementById('lista').innerHTML = html || '<p style="color:#999;font-size:.85rem">Sin productos.</p>';
}

function toggle(id) {
  if (seleccionados.has(id)) seleccionados.delete(id);
  else seleccionados.add(id);
  var el = document.getElementById('pi-' + id);
  if (el) {
    el.classList.toggle('sel', seleccionados.has(id));
    var cb = el.querySelector('input[type=checkbox]');
    if (cb) cb.checked = seleccionados.has(id);
  }
  actualizarContador();
}

function toggleTodos(val) {
  if (val) productos.forEach(function(p){ seleccionados.add(p.id); });
  else seleccionados.clear();
  renderLista();
  actualizarContador();
}

function actualizarContador() {
  document.getElementById('contador').textContent = seleccionados.size + ' de ' + productos.length + ' seleccionados';
}

function esc(s) {
  return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}
async function eliminarUltima() {
  var ids = Array.from(seleccionados);
  if (!ids.length) { alert('Selecciona al menos un producto.'); return; }
  if (!confirm('Eliminar la ultima imagen de ' + ids.length + ' productos?')) return;
  document.getElementById('btn-eliminar').disabled = true;
  document.getElementById('barra-cont2').style.display = 'block';
  var log = document.getElementById('log2');
  log.innerHTML = '';
  var ok = 0;
  var errores = 0;
  for (var i = 0; i < ids.length; i++) {
    var pid = ids[i];
    var prod = productos.find(function(p){ return p.id === pid; });
    document.getElementById('progreso2').style.width = Math.round(((i+1)/ids.length)*100) + '%';
    log.innerHTML += '<div>Eliminando de &quot;' + esc(prod ? prod.nombre : pid) + '&quot;... ';
    try {
      var r = await fetch('/api/eliminar-ultima-imagen', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({product_ids: [pid]})
      });
      var data = await r.json();
      if (data.ok > 0) { ok++; log.innerHTML += '<span style="color:#16a34a">OK</span></div>'; }
      else { errores++; log.innerHTML += '<span style="color:#dc2626">Error</span></div>'; }
    } catch(e) {
      errores++;
      log.innerHTML += '<span style="color:#dc2626">Error de red</span></div>';
    }
    log.scrollTop = log.scrollHeight;
    await new Promise(function(res){ setTimeout(res, 600); });
  }
  log.innerHTML += '<div style="font-weight:600;margin-top:.5rem">Listo: ' + ok + ' OK, ' + errores + ' errores</div>';
  document.getElementById('btn-eliminar').disabled = false;
}
async function subir() {
  if (!imgBase64) { alert('Primero subi una imagen.'); return; }
  var ids = Array.from(seleccionados);
  if (!ids.length) { alert('Selecciona al menos un producto.'); return; }
  if (!confirm('Subir imagen a ' + ids.length + ' productos?')) return;
  document.getElementById('btn-subir').disabled = true;
  document.getElementById('barra-cont').style.display = 'block';
  var log = document.getElementById('log');
  log.innerHTML = '';
  var ok = 0;
  var errores = 0;
  for (var i = 0; i < ids.length; i++) {
    var pid = ids[i];
    var prod = productos.find(function(p){ return p.id === pid; });
    var pct = Math.round(((i+1) / ids.length) * 100);
    document.getElementById('progreso').style.width = pct + '%';
    log.innerHTML += '<div>Subiendo a &quot;' + esc(prod ? prod.nombre : pid) + '&quot;... ';
    try {
      var r = await fetch('/api/subir-tabla-talles', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({product_ids: [pid], filename: imgFilename, attachment: imgBase64})
      });
      var data = await r.json();
      if (data.ok > 0) { ok++; log.innerHTML += '<span style="color:#16a34a">OK</span></div>'; }
      else { errores++; log.innerHTML += '<span style="color:#dc2626">Error: ' + esc(JSON.stringify(data.errores)) + '</span></div>'; }
    } catch(e) {
      errores++;
      log.innerHTML += '<span style="color:#dc2626">Error de red</span></div>';
    }
    log.scrollTop = log.scrollHeight;
    await new Promise(function(res){ setTimeout(res, 600); });
  }
  log.innerHTML += '<div style="font-weight:600;margin-top:.5rem">Listo: ' + ok + ' OK, ' + errores + ' errores</div>';
  document.getElementById('btn-subir').disabled = false;
}
</script>
</body>
</html>"""

@app.get("/tabla-talles", response_class=HTMLResponse)
def tabla_talles():
    return TABLA_TALLES_HTML
